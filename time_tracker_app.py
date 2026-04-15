"""
RTL Time Tracker — Streamlit Web App
=====================================
Reads from and writes back to individual staff Excel tracking files.

To run locally:
    pip install -r requirements.txt
    streamlit run time_tracker_app.py

Put this file in the same folder as the *_2026TimeTracking.xlsx files.
"""

import streamlit as st
import pandas as pd
import openpyxl
import requests
import base64
import io
from pathlib import Path
from datetime import datetime, date, timedelta
import calendar
import plotly.graph_objects as go

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="RTL KPI Tracker",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════════════
# CVU BRAND TOKENS  (from Brand Guidelines V3.0 – dark theme)
# ══════════════════════════════════════════════════════════════════════════════

CVU_BLACK    = "#171717"   # Primary background
CVU_SURFACE  = "#282828"   # Card / chart background
CVU_BORDER   = "#4F4F4F"   # Grid lines, dividers
CVU_WHITE    = "#FCFCFC"   # Primary text
CVU_GRAY     = "#9E9E9E"   # Secondary text / axis labels
CVU_GREEN    = "#B4E817"   # Volt Green – primary accent

# Ordered accent palette for chart series (Indigo, Solar, Aqua, Teal, Plum, Ember)
CVU_PALETTE = [
    "#516BFF",
    "#FF9F18",
    "#54D9E7",
    "#34C684",
    "#C63AD2",
    "#FA3F26",
]

# ── Minimal global styles ─────────────────────────────────────────────────────
# Page background, text, and widget colours come from .streamlit/config.toml.
# Only inject what config.toml can't express: the Inter web font, metric card
# accent border, and the category-header chip used in daily entry.
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, .stApp {
    font-family: 'Inter', Arial, sans-serif !important;
}

/* Metric card — dark card background + left accent stripe */
div[data-testid="stMetric"] {
    background-color: #282828 !important;
    border-radius: 6px;
    padding: 12px 16px;
    border-left: 3px solid var(--primary-color, #B4E817);
}

/* Category section header chip in daily entry */
.cat-header {
    border-left: 3px solid var(--primary-color, #B4E817);
    padding: 6px 12px;
    border-radius: 0 4px 4px 0;
    font-weight: 600;
    font-size: 0.88rem;
    letter-spacing: 0.03em;
    margin-top: 14px;
    margin-bottom: 4px;
}

/* Primary buttons: dark text on #B4E817 lime-green background for legibility */
button[kind="primary"],
div[data-testid="stButton"] > button[kind="primary"],
div[data-testid="stFormSubmitButton"] > button[kind="primary"] {
    color: #171717 !important;
}

/* Multiselect tag bubbles: dark text on lime-green background */
span[data-baseweb="tag"] {
    color: #171717 !important;
}
span[data-baseweb="tag"] span,
span[data-baseweb="tag"] svg {
    color: #171717 !important;
    fill: #171717 !important;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

# When running locally, set DATA_DIR in .streamlit/secrets.toml to point at
# your OneDrive folder so reads and writes go to the right place.
# On Streamlit Cloud this falls back to the repo directory (files come from GitHub).
try:
    _candidate = Path(st.secrets["DATA_DIR"])
    # Only use the secret path if it actually exists on this machine.
    # On Streamlit Cloud the OneDrive path won't exist, so it falls back
    # to the repo directory where the Excel files are committed.
    DATA_DIR = _candidate if _candidate.exists() else Path(__file__).parent
except Exception:
    DATA_DIR = Path(__file__).parent

STAFF = {
    "D. Safarik":  "DSafarik_2026TimeTracking.xlsx",
    "I. Work":     "IWork_2026TimeTracking.xlsx",
    "S. Ursini":   "SUrsini_2026TimeTracking.xlsx",
    "W. Miranda":  "WMiranda_2026TimeTracking.xlsx",
}
STAFF = {name: DATA_DIR / fname for name, fname in STAFF.items()}

TRACKER_SHEET  = "Tracking"
DATE_ROW       = 5   # Row with actual datetime values (row 4 has =C5 formulas)
DATA_START_ROW = 7   # First row with task data
DATE_START_COL = 3   # Column C = first date column

GITHUB_OWNER  = "sisaacwork"
GITHUB_REPO   = "rtl-time-tracker"
GITHUB_BRANCH = "main"

# Codes to treat as absences — excluded from the 900s/other and funded/unfunded charts
ABSENCE_CODES = {"120", "121", "122", "123", "123", "124"}

# Sub-codes that count as externally funded work
FUNDED_CODES = {
    "903b", "903c",
    "904a", "904b",
    "905a", "905b",
    "906a", "906b",
    "910", "915", "916", "917",
}

# ══════════════════════════════════════════════════════════════════════════════
# FINANCIAL TRACKING CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

FINANCES_FILE = "finances.xlsx"

INCOME_STATUSES = ["Paid", "Invoiced", "Contracted", "Verbal", "Pipeline"]

# Base accounting codes — all staff costs are bundled under 903.
# Users can add custom codes inside the app.
ACCOUNTING_CODES = {
    "901a": "Venice Research Office",
    "901b": "Canada Research Office",
    "902":  "Seed Funding",
    "903":  "Short-term Projects",
    "904":  "Cities for People",
    "905":  "Sustainability Program",
    "906":  "T+U Innovation",
    "907":  "Megatalls Assembly",
    "910":  "ClimateWorks Code Research",
    "917":  "Commissioned Research",
}

STATUS_COLORS = {
    "Paid":       "#B4E817",   # CVU_GREEN
    "Invoiced":   "#516BFF",   # CVU_PALETTE[0]
    "Contracted": "#54D9E7",   # CVU_PALETTE[2]
    "Verbal":     "#FF9F18",   # CVU_PALETTE[1]
    "Pipeline":   "#4F4F4F",   # CVU_BORDER (muted)
}

# ══════════════════════════════════════════════════════════════════════════════
# CONTENT KPI CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

CONTENT_FILE = "content_projects.xlsx"

CONTENT_TYPES = ["Book", "White paper", "Data study", "Magazine", "PDF report"]

DEFAULT_PILLARS = [
    "Program Partnership",
    "Committee-Driven",
    "Commissioned Research",
    "Conference Sponsorship",
    "Advertising and Sales-Driven",
    "Other Sponsorship",
]

PILLAR_COLORS = {
    "Program Partnership":          "#516BFF",
    "Committee-Driven":             "#C63AD2",
    "Commissioned Research":        "#FF9F18",
    "Conference Sponsorship":       "#54D9E7",
    "Advertising and Sales-Driven": "#FA3F26",
    "Other Sponsorship":            "#34C684",
}

DEFAULT_PILLAR_COLOR = "#9E9E9E"  # fallback for any custom pillars added by user

CONTENT_ACCT_CODES = {
    "904a": "City Advocacy",
    "905a": "Sustainability Program",
    "906a": "T+U Innovation",
    "910":  "Climateworks Code Research",
    "916":  "Commissioned Research",
    "701":  "Vertical Urbanism Magazine",
    "702":  "Tall + Urban Awards Book",
    "703":  "Conference Publications, Proceedings & Reports",
    "704":  "Research Reports",
    "705":  "Technical Guides",
    "710":  "Other Publications / Data Handbook",
}

RTL_OWNERS         = ["DS", "IW", "RD", "SU", "WM"]
SPONSORSHIP_TYPES  = ["Program Partnership", "Commissioned Research", "VU Advert", "Other"]
PROJECT_STATUSES   = [
    "Obligatory sponsor deliverable",
    "Internal commitment",
    "Internal idea",
    "Project completed",
]
FORMAT_TYPES       = ["Digital", "Print", "Both", "TBD"]
CONTENT_GENERATORS = [
    "RTL",
    "Client + RTL",
    "Other CVU Team",
    "Conference Speakers",
    "Committee",
    "External Contributors",
]
FUNDING_SOURCES    = ["Conference Sponsor", "Program Partner", "Commissioned Research", "Sales", "None"]

# Column order for the projects sheet — do not reorder without a migration
PROJECTS_COLS = [
    "id", "title", "type", "pillar", "acct_code", "owner",
    "sponsored", "sponsorship_type", "sponsorship_other",
    "status", "format", "content_generator", "committee_name",
    "funding_source", "program_partner_name",
    "budget", "est_hours",
    "draft_delivered", "draft_commented", "draft_completed",
    "layout1_delivered", "layout1_commented", "layout2_delivered", "layout2_approved",
    "print_date", "go_live_date",
    "notes", "pct_override", "confirmed_pending",
]

# ══════════════════════════════════════════════════════════════════════════════
# GITHUB / CLOUD HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _github_token():
    """Return the GitHub token from Streamlit secrets, or None if running locally."""
    try:
        return st.secrets["GITHUB_TOKEN"]
    except Exception:
        return None


def _is_cloud() -> bool:
    """True when running on Streamlit Community Cloud (token available)."""
    return _github_token() is not None


def _github_commit(filename: str, content_bytes: bytes) -> tuple:
    """
    Upload (or update) a file in the GitHub repo.
    Returns (success: bool, message: str).
    """
    token = _github_token()
    if not token:
        return False, "No GitHub token configured."

    url = (
        f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}"
        f"/contents/{filename}"
    )
    headers = {
        "Authorization": f"token {token}",
        "Accept":        "application/vnd.github.v3+json",
    }
    encoded = base64.b64encode(content_bytes).decode()
    commit_msg = (
        f"Update: {filename} "
        f"[{datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC]"
    )

    def _fetch_sha():
        # Always pin to the target branch so the SHA matches exactly.
        r = requests.get(f"{url}?ref={GITHUB_BRANCH}", headers=headers)
        return r.json().get("sha") if r.status_code == 200 else None

    def _do_put(sha):
        payload = {"message": commit_msg, "content": encoded, "branch": GITHUB_BRANCH}
        if sha:
            payload["sha"] = sha
        return requests.put(url, headers=headers, json=payload)

    resp = _do_put(_fetch_sha())

    # If we get a 409 (SHA conflict), re-fetch and retry once.
    if resp.status_code == 409:
        resp = _do_put(_fetch_sha())

    if resp.status_code in (200, 201):
        return True, "Saved to GitHub."
    return False, f"GitHub error {resp.status_code}: {resp.text[:200]}"


# ══════════════════════════════════════════════════════════════════════════════
# FINANCES DATA — load / save
# ══════════════════════════════════════════════════════════════════════════════

TXNS_COLS = ["id", "date", "type", "amount", "code", "code_name",
             "description", "status", "notes"]


def _empty_transactions() -> pd.DataFrame:
    return pd.DataFrame(columns=TXNS_COLS)


def _fetch_finances_bytes():
    """
    Return raw bytes of finances.xlsx from GitHub (cloud) or disk (local).
    Returns None if the file doesn't exist yet.
    """
    if _is_cloud():
        url = (
            f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}"
            f"/contents/{FINANCES_FILE}?ref={GITHUB_BRANCH}"
        )
        resp = requests.get(url, headers={
            "Authorization": f"token {_github_token()}",
            "Accept":        "application/vnd.github.v3+json",
        })
        if resp.status_code == 404:
            return None
        if resp.status_code != 200:
            raise RuntimeError(f"GitHub error fetching finances: {resp.status_code}")
        return base64.b64decode(resp.json()["content"].replace("\n", ""))
    else:
        local = DATA_DIR / FINANCES_FILE
        if not local.exists():
            return None
        return local.read_bytes()


@st.cache_data(ttl=30, show_spinner="Loading financial data...")
def load_finances():
    """
    Returns (transactions_df, settings_dict, all_codes_dict).
    Creates empty structures if the file doesn't exist yet.
    """
    raw = _fetch_finances_bytes()

    if raw is None:
        return _empty_transactions(), {}, dict(ACCOUNTING_CODES)

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # --- transactions sheet ---
    records = []
    if "transactions" in wb.sheetnames:
        ws = wb["transactions"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            records.append(dict(zip(TXNS_COLS, row)))

    txns = pd.DataFrame(records, columns=TXNS_COLS) if records else _empty_transactions()
    if not txns.empty:
        txns["date"]   = pd.to_datetime(txns["date"])
        txns["amount"] = pd.to_numeric(txns["amount"], errors="coerce").fillna(0.0)

    # --- settings sheet ---
    settings     = {}
    custom_codes = {}
    if "settings" in wb.sheetnames:
        ws = wb["settings"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] is None:
                continue
            key = str(row[0])
            val = row[1]
            if key.startswith("custom_code_"):
                code = key[len("custom_code_"):]
                custom_codes[code] = str(val) if val else code
            else:
                try:
                    settings[key] = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    settings[key] = val

    all_codes = dict(ACCOUNTING_CODES)
    all_codes.update(custom_codes)

    return txns, settings, all_codes


def save_finances(txns: pd.DataFrame, settings: dict, custom_codes: dict) -> tuple:
    """
    Persist transactions + settings to finances.xlsx and commit to GitHub.
    Returns (success: bool, message: str).
    """
    wb    = openpyxl.Workbook()
    ws_t  = wb.active
    ws_t.title = "transactions"
    ws_t.append(TXNS_COLS)

    for _, row in txns.iterrows():
        ws_t.append([
            int(row.get("id", 0)),
            row["date"].date() if pd.notna(row.get("date")) else None,
            str(row.get("type", "")),
            float(row.get("amount", 0)),
            str(row.get("code", "")),
            str(row.get("code_name", "")),
            str(row.get("description", "")),
            str(row.get("status", "")),
            str(row.get("notes", "")),
        ])

    ws_s = wb.create_sheet("settings")
    for key, val in settings.items():
        ws_s.append([key, val])
    for code, name in custom_codes.items():
        ws_s.append([f"custom_code_{code}", name])

    buf = io.BytesIO()
    wb.save(buf)
    content_bytes = buf.getvalue()

    local_path = DATA_DIR / FINANCES_FILE
    if _is_cloud():
        ok, msg = _github_commit(FINANCES_FILE, content_bytes)
        if ok:
            local_path.write_bytes(content_bytes)
    else:
        local_path.write_bytes(content_bytes)
        ok, msg = True, "Saved."

    load_finances.clear()
    return ok, msg


# ══════════════════════════════════════════════════════════════════════════════
# CONTENT PROJECTS DATA — load / save
# ══════════════════════════════════════════════════════════════════════════════

def _empty_projects() -> pd.DataFrame:
    return pd.DataFrame(columns=PROJECTS_COLS)


def _fetch_content_bytes():
    """Return raw bytes of content_projects.xlsx from GitHub (cloud) or disk (local)."""
    if _is_cloud():
        url = (
            f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}"
            f"/contents/{CONTENT_FILE}?ref={GITHUB_BRANCH}"
        )
        resp = requests.get(url, headers={
            "Authorization": f"token {_github_token()}",
            "Accept":        "application/vnd.github.v3+json",
        })
        if resp.status_code == 404:
            return None
        if resp.status_code != 200:
            raise RuntimeError(f"GitHub error fetching content projects: {resp.status_code}")
        return base64.b64decode(resp.json()["content"].replace("\n", ""))
    else:
        local = DATA_DIR / CONTENT_FILE
        if not local.exists():
            return None
        return local.read_bytes()


@st.cache_data(ttl=3600, show_spinner=False)
def load_content_projects():
    """
    Returns (projects_df, settings_dict, all_pillars_list, all_codes_dict).
    Creates empty structures if the file doesn't exist yet.
    """
    raw = _fetch_content_bytes()
    if raw is None:
        return _empty_projects(), {}, list(DEFAULT_PILLARS), dict(CONTENT_ACCT_CODES)

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # --- projects sheet ---
    records = []
    if "projects" in wb.sheetnames:
        ws = wb["projects"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            records.append(dict(zip(PROJECTS_COLS, row)))

    projects = (pd.DataFrame(records, columns=PROJECTS_COLS)
                if records else _empty_projects())
    if not projects.empty:
        for col in ("budget", "est_hours"):
            projects[col] = pd.to_numeric(projects[col], errors="coerce").fillna(0.0)

    # --- settings sheet ---
    settings       = {}
    custom_pillars = []
    custom_codes   = {}
    if "settings" in wb.sheetnames:
        ws = wb["settings"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] is None:
                continue
            key = str(row[0])
            val = row[1]
            if key.startswith("custom_pillar_"):
                if val:
                    custom_pillars.append(str(val))
            elif key.startswith("custom_content_code_"):
                code = key[len("custom_content_code_"):]
                custom_codes[code] = str(val) if val else code
            else:
                try:
                    settings[key] = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    settings[key] = val

    all_pillars = list(DEFAULT_PILLARS) + [p for p in custom_pillars
                                           if p not in DEFAULT_PILLARS]
    all_codes   = dict(CONTENT_ACCT_CODES)
    all_codes.update(custom_codes)

    return projects, settings, all_pillars, all_codes


def save_content_projects(
    projects: pd.DataFrame,
    settings: dict,
    custom_pillars: list,
    custom_codes: dict,
) -> tuple:
    """Persist projects + settings to content_projects.xlsx and commit to GitHub."""
    wb   = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = "projects"
    ws_p.append(PROJECTS_COLS)

    for _, row in projects.iterrows():
        ws_p.append([
            int(row.get("id", 0)),
            str(row.get("title",               "") or ""),
            str(row.get("type",                "") or ""),
            str(row.get("pillar",              "") or ""),
            str(row.get("acct_code",           "") or ""),
            str(row.get("owner",               "") or ""),
            str(row.get("sponsored",           "No")),
            str(row.get("sponsorship_type",    "") or ""),
            str(row.get("sponsorship_other",   "") or ""),
            str(row.get("status",              "") or ""),
            str(row.get("format",              "") or ""),
            str(row.get("content_generator",   "") or ""),
            str(row.get("committee_name",      "") or ""),
            str(row.get("funding_source",      "") or ""),
            str(row.get("program_partner_name","") or ""),
            float(row.get("budget",     0.0) or 0.0),
            float(row.get("est_hours",  0.0) or 0.0),
            str(row.get("draft_delivered",    "") or ""),
            str(row.get("draft_commented",    "") or ""),
            str(row.get("draft_completed",    "") or ""),
            str(row.get("layout1_delivered",  "") or ""),
            str(row.get("layout1_commented",  "") or ""),
            str(row.get("layout2_delivered",  "") or ""),
            str(row.get("layout2_approved",   "") or ""),
            str(row.get("print_date",         "") or ""),
            str(row.get("go_live_date",       "") or ""),
            str(row.get("notes",              "") or ""),
            float(row.get("pct_override", 0.0) or 0.0),
            str(row.get("confirmed_pending", "Confirmed") or "Confirmed"),
        ])

    ws_s = wb.create_sheet("settings")
    for key, val in settings.items():
        ws_s.append([key, val])
    for i, pillar in enumerate(custom_pillars):
        ws_s.append([f"custom_pillar_{i}", pillar])
    for code, name in custom_codes.items():
        ws_s.append([f"custom_content_code_{code}", name])

    buf = io.BytesIO()
    wb.save(buf)
    content_bytes = buf.getvalue()

    local_path = DATA_DIR / CONTENT_FILE
    if _is_cloud():
        ok, msg = _github_commit(CONTENT_FILE, content_bytes)
        if ok:
            local_path.write_bytes(content_bytes)
    else:
        local_path.write_bytes(content_bytes)
        ok, msg = True, "Saved."

    load_content_projects.clear()
    return ok, msg


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL PARSING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _date_cols(ws) -> dict:
    """Return {date_object: col_index} for every date column in the sheet."""
    out = {}
    for c in range(DATE_START_COL, ws.max_column + 1):
        v = ws.cell(row=DATE_ROW, column=c).value
        if isinstance(v, datetime):
            out[v.date()] = c
    return out


def _task_rows(ws) -> list:
    """Return [(task_name, row_index), ...] for every task row in the sheet."""
    out = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and v.strip():
            out.append((v.strip(), r))
    return out


def is_child(task: str) -> bool:
    """Child tasks start with '- '."""
    return task.startswith("- ")


def short_name(task: str) -> str:
    """Strip the '- ' prefix for cleaner display."""
    return task[2:].strip() if is_child(task) else task


def category_code(task: str) -> str:
    """Map any task to its top-level 3-digit code (e.g. '- 901a:...' → '900')."""
    t = task.lstrip("- ").strip()
    if t and t[0].isdigit():
        return t[0] + "00"
    return "Other"


def task_subcode(task: str) -> str:
    """
    Extract the full alphanumeric sub-code from a task name.
    e.g. '- 903b: CTBUHx Chicago' → '903b'
         '- 120: Annual Leave'    → '120'
    Returns an empty string if no leading code is found.
    """
    import re
    t = task.lstrip("- ").strip()
    m = re.match(r"^(\d+[a-z]*)", t, re.IGNORECASE)
    return m.group(1).lower() if m else ""


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING  (cached 30 s — Refresh button busts manually)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=30, show_spinner="Loading tracker data...")
def load_all() -> pd.DataFrame:
    """
    Read every staff Excel file → single long-format DataFrame:
        person | date | task | hours | week | month | quarter | code | category
    Only rows with hours > 0 are included.
    """
    records = []
    for person, path in STAFF.items():
        if not _is_cloud() and not path.exists():
            st.warning(f"File not found, skipping: {path.name}")
            continue
        try:
            wb = _fetch_workbook(person, data_only=True)
        except Exception as e:
            st.warning(f"Could not load {path.name}: {e}")
            continue
        ws = wb[TRACKER_SHEET]
        dc = _date_cols(ws)
        tr = _task_rows(ws)
        for task, r in tr:
            for d, c in dc.items():
                h = ws.cell(row=r, column=c).value
                if h:
                    records.append({
                        "person": person,
                        "date":   d,
                        "task":   task,
                        "hours":  float(h),
                    })

    if not records:
        return pd.DataFrame(columns=["person", "date", "task", "hours",
                                      "week", "month", "quarter", "code", "category"])

    df = pd.DataFrame(records)
    df["date"]    = pd.to_datetime(df["date"])
    df["week"]    = df["date"].dt.to_period("W").astype(str)
    df["month"]   = df["date"].dt.to_period("M").astype(str)
    df["quarter"] = df["date"].dt.to_period("Q").astype(str)
    df["code"]    = df["task"].apply(category_code)

    # Build code → full parent category label (e.g. "900" → "900 Research & Thought Leadership")
    code_names = {}
    for task in df["task"].unique():
        if not is_child(task):
            code_names[category_code(task)] = task
    df["category"] = df["code"].map(code_names).fillna(df["code"])

    return df


# ══════════════════════════════════════════════════════════════════════════════
# PER-PERSON DATA ACCESS
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_workbook(person: str, data_only: bool = True):
    """
    Load a person's workbook.
    On Streamlit Cloud: fetches the latest bytes directly from GitHub API
    so the app always reflects the current state of the repo, not a stale
    snapshot from deploy time.
    Locally: reads straight from disk as before.
    """
    filename = STAFF[person].name
    if _is_cloud():
        url = (
            f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}"
            f"/contents/{filename}?ref={GITHUB_BRANCH}"
        )
        resp = requests.get(url, headers={
            "Authorization": f"token {_github_token()}",
            "Accept":        "application/vnd.github.v3+json",
        })
        if resp.status_code != 200:
            raise FileNotFoundError(f"Could not fetch {filename} from GitHub: {resp.status_code}")
        raw = base64.b64decode(resp.json()["content"].replace("\n", ""))
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=data_only)
    else:
        return openpyxl.load_workbook(STAFF[person], data_only=data_only)


def workdays(person: str) -> list:
    """Sorted list of all tracked working dates for this person."""
    wb = _fetch_workbook(person, data_only=True)
    return sorted(_date_cols(wb[TRACKER_SHEET]).keys())


def hours_on_date(person: str, d: date) -> dict:
    """Return {task_name: hours} for a given person + date (0 if not entered)."""
    wb = _fetch_workbook(person, data_only=True)
    ws = wb[TRACKER_SHEET]
    dc = _date_cols(ws)
    c  = dc.get(d)
    if c is None:
        return {}
    return {
        t: float(ws.cell(row=r, column=c).value or 0)
        for t, r in _task_rows(ws)
    }


def task_structure(person: str) -> list:
    """Return [(task_name, is_child_bool), ...] in spreadsheet order."""
    wb = _fetch_workbook(person, data_only=True)
    ws = wb[TRACKER_SHEET]
    return [(t, is_child(t)) for t, _ in _task_rows(ws)]


def save_hours(person: str, d: date, hours: dict) -> tuple:
    """
    Write updated hours back to the person's Excel file.

    Two-step approach:
      1. Open data_only=True to resolve column index for the target date.
      2. Open without data_only (preserving formulas) to write the values.

    If running on Streamlit Cloud, commits the result to GitHub instead of
    writing directly to disk.

    Returns (success: bool, message: str).
    """
    path = STAFF[person]

    # Step 1: resolve date → column (data_only=True to get computed date values)
    wb_r = _fetch_workbook(person, data_only=True)
    dc   = _date_cols(wb_r[TRACKER_SHEET])
    c    = dc.get(d)
    if c is None:
        return False, f"Date {d} not found in the tracker spreadsheet."

    # Step 2: open without data_only so formulas stay intact
    wb = _fetch_workbook(person, data_only=False)
    ws = wb[TRACKER_SHEET]
    for t, r in _task_rows(ws):
        if t in hours:
            v = hours[t]
            ws.cell(row=r, column=c).value = v if v > 0 else None

    if _is_cloud():
        buf = io.BytesIO()
        wb.save(buf)
        content_bytes = buf.getvalue()
        ok, msg = _github_commit(path.name, content_bytes)
        if ok:
            # Also write to the local snapshot so load_all sees fresh data
            # without waiting for a GitHub API round-trip on the next read.
            path.write_bytes(content_bytes)
    else:
        wb.save(path)
        ok, msg = True, "Saved to Excel."

    load_all.clear()
    return ok, msg


# ══════════════════════════════════════════════════════════════════════════════
# CHART THEME HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _chart_colors() -> dict:
    """
    Return a dict of colour tokens that adapt to the current dark/light setting.
    Reads st.session_state["dark_mode"] — defaults to True (dark).
    """
    dark = st.session_state.get("dark_mode", True)
    if dark:
        return {
            "text":     CVU_WHITE,      # primary text / tick labels
            "subtext":  CVU_GRAY,       # secondary / axis labels
            "plot_bg":  CVU_BLACK,      # chart plot area background
            "grid":     CVU_BORDER,     # grid lines
            "hover_bg": CVU_SURFACE,    # tooltip background
            "refline":  CVU_WHITE,      # reference / annotation lines
        }
    else:
        return {
            "text":     "#1A1A1A",
            "subtext":  "#555555",
            "plot_bg":  "rgba(0,0,0,0)",   # transparent — page bg shows through
            "grid":     "#CCCCCC",
            "hover_bg": "#EBEBEB",
            "refline":  "#555555",
        }


def _chart_base(**overrides) -> dict:
    """
    Base Plotly layout dict with CVU styling that adapts to dark/light mode.
    Pass keyword args to override any key.
    """
    c = _chart_colors()
    layout = dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor=c["plot_bg"],
        font=dict(family="Inter, Arial, sans-serif", color=c["text"], size=12),
        xaxis=dict(
            gridcolor=c["grid"],
            zerolinecolor=c["grid"],
            tickfont=dict(color=c["subtext"], size=11),
        ),
        yaxis=dict(
            gridcolor=c["grid"],
            zerolinecolor=c["grid"],
            tickfont=dict(color=c["subtext"], size=11),
            automargin=True,
        ),
        legend=dict(
            bgcolor="rgba(0,0,0,0)",
            font=dict(color=c["text"], size=11),
        ),
        margin=dict(l=10, r=20, t=36, b=30),
        hoverlabel=dict(
            bgcolor=c["hover_bg"],
            font=dict(color=c["text"], family="Inter, Arial, sans-serif"),
        ),
    )
    layout.update(overrides)
    return layout


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: DAILY ENTRY
# ══════════════════════════════════════════════════════════════════════════════

def view_daily_entry(person: str):
    st.subheader("Daily Time Entry")

    days = workdays(person)
    if not days:
        st.error("No tracked dates found in this person's file.")
        return

    today   = date.today()
    default = today if today in days else days[-1]

    selected = st.date_input(
        "Select date",
        value=default,
        min_value=days[0],
        max_value=days[-1],
    )

    if selected not in days:
        st.warning("That date is not a tracked workday in the spreadsheet.")
        return

    current   = hours_on_date(person, selected)
    structure = task_structure(person)

    # Group tasks into (parent_category, [(child_task, current_hours), ...])
    groups       = []
    cur_parent   = None
    cur_children = []
    for t, child in structure:
        if not child:
            if cur_parent is not None:
                groups.append((cur_parent, cur_children))
            cur_parent   = t
            cur_children = []
        else:
            cur_children.append((t, current.get(t, 0.0)))
    if cur_parent is not None:
        groups.append((cur_parent, cur_children))

    st.caption(
        f"**{selected.strftime('%A, %B %d, %Y')}** — "
        "enter hours in 0.25 increments (0.25 = 15 min, 1.0 = 1 hr)"
    )

    # Inputs live outside a form so the total updates on every change
    new_vals = {}
    for parent, kids in groups:
        if not kids:
            continue
        st.markdown(
            f"<div class='cat-header'>{parent}</div>",
            unsafe_allow_html=True,
        )
        for task, cur_h in kids:
            label = short_name(task)
            left, right = st.columns([5, 1])
            left.markdown(f"&nbsp;&nbsp;&nbsp;{label}")
            v = right.number_input(
                label, label_visibility="collapsed",
                min_value=0.0, max_value=24.0, step=0.25,
                value=float(cur_h),
                key=f"de_{task}_{selected}",
            )
            new_vals[task] = v

    total    = sum(new_vals.values())
    overtime = max(0.0, total - 8.0)
    st.divider()
    tcol1, tcol2 = st.columns(2)
    tcol1.metric("Daily Total", f"{total:.2f} h")
    tcol2.metric("Overtime",    f"{overtime:.2f} h", delta_color="inverse")

    if st.button("Save Entry", type="primary", use_container_width=True):
        ok, msg = save_hours(person, selected, new_vals)
        if ok:
            st.toast("Entry saved successfully.")
            st.rerun()
        else:
            st.error(msg)


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: BULK EDIT
# ══════════════════════════════════════════════════════════════════════════════

def view_bulk_edit(person: str):
    st.subheader("Bulk Edit")

    days = workdays(person)
    if not days:
        st.error("No tracked dates found.")
        return

    today = date.today()
    c1, c2 = st.columns(2)
    d_from = c1.date_input("From", value=max(days[0],  today - timedelta(days=13)))
    d_to   = c2.date_input("To",   value=min(days[-1], today))

    in_range = [d for d in days if d_from <= d <= d_to]
    if not in_range:
        st.info("No tracked workdays in that range.")
        return

    structure = task_structure(person)
    parents   = [t for t, ch in structure if not ch]
    sel_cats  = st.multiselect(
        "Filter by category (select which to show)",
        options=parents,
        default=parents,
    )

    allowed = set()
    cur_p   = None
    for t, ch in structure:
        if not ch:
            cur_p = t
        elif cur_p in sel_cats:
            allowed.add(t)

    child_tasks = [t for t, ch in structure if ch and t in allowed]
    if not child_tasks:
        st.info("No tasks to show for the selected categories.")
        return

    display_names = [short_name(t) for t in child_tasks]
    name_map      = dict(zip(display_names, child_tasks))

    rows        = []
    date_labels = []
    for d in in_range:
        dh    = hours_on_date(person, d)
        label = d.strftime("%Y-%m-%d (%a)")
        date_labels.append(label)
        row   = {dn: dh.get(ft, 0.0) for dn, ft in name_map.items()}
        rows.append(row)

    df_edit = pd.DataFrame(rows, index=date_labels)
    df_edit.index.name = "Date"

    st.caption("Edit cells directly. Click Save All Changes when done.")
    edited = st.data_editor(df_edit, use_container_width=True, num_rows="fixed")

    if st.button("Save All Changes", type="primary", use_container_width=True):
        saved  = 0
        errors = []
        for d, label in zip(in_range, date_labels):
            row     = edited.loc[label]
            updates = {
                name_map[col]: float(row[col])
                for col in row.index if col in name_map
            }
            ok, msg = save_hours(person, d, updates)
            if ok:
                saved += 1
            else:
                errors.append(msg)

        if errors:
            st.error("\n".join(errors))
        else:
            st.success(f"Saved {saved} of {len(in_range)} days.")
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: HISTORY (personal time history)
# ══════════════════════════════════════════════════════════════════════════════

def view_history(person: str, df: pd.DataFrame):
    st.subheader("My History")
    c = _chart_colors()

    pdf = df[(df["person"] == person) & df["task"].str.startswith("- ")]
    if pdf.empty:
        st.info("No time entries found yet.")
        return

    # Date range filter
    min_d = pdf["date"].min().date()
    max_d = pdf["date"].max().date()
    c1, c2 = st.columns(2)
    d_from = c1.date_input("From", value=min_d, key="hist_from")
    d_to   = c2.date_input("To",   value=max_d, key="hist_to")

    pdf = pdf[(pdf["date"].dt.date >= d_from) & (pdf["date"].dt.date <= d_to)]
    if pdf.empty:
        st.info("No entries in that date range.")
        return

    # KPI summary
    daily_totals = pdf.groupby("date")["hours"].sum()
    k1, k2, k3 = st.columns(3)
    k1.metric("Total Hours",      f"{pdf['hours'].sum():,.1f}")
    k2.metric("Days Worked",      str(pdf["date"].dt.date.nunique()))
    k3.metric("Avg Hours / Day",  f"{daily_totals.mean():.1f}")
    st.divider()

    # Chart 1: Hours by category
    cat_df = (
        pdf.groupby("category")["hours"]
           .sum()
           .reset_index()
           .sort_values("hours")
    )
    fig_cat = go.Figure(go.Bar(
        y=cat_df["category"],
        x=cat_df["hours"],
        orientation="h",
        marker_color=CVU_GREEN,
        hovertemplate="%{y}<br>Hours: %{x:.1f}<extra></extra>",
    ))
    fig_cat.update_layout(**_chart_base(
        title=dict(text="Hours by Category", font=dict(color=c["text"], size=14)),
        height=max(300, len(cat_df) * 55),
        yaxis=dict(automargin=True, tickfont=dict(color=c["text"], size=11),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        xaxis=dict(title="Hours", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
    ))
    st.plotly_chart(fig_cat, use_container_width=True)

    # Chart 2: Daily hours + cumulative line
    daily = pdf.groupby("date")["hours"].sum().reset_index().sort_values("date")
    daily["cumulative"] = daily["hours"].cumsum()

    fig_line = go.Figure()
    fig_line.add_trace(go.Bar(
        x=daily["date"],
        y=daily["hours"],
        name="Daily Hours",
        marker_color=CVU_GREEN,
        opacity=0.75,
        hovertemplate="Date: %{x|%b %d}<br>Hours: %{y:.2f}<extra></extra>",
    ))
    fig_line.add_trace(go.Scatter(
        x=daily["date"],
        y=daily["cumulative"],
        name="Cumulative",
        line=dict(color=CVU_PALETTE[0], width=2),
        yaxis="y2",
        hovertemplate="Date: %{x|%b %d}<br>Cumulative: %{y:.1f} hrs<extra></extra>",
    ))
    fig_line.update_layout(**_chart_base(
        title=dict(text="Daily Hours & Running Total", font=dict(color=c["text"], size=14)),
        height=320,
        yaxis=dict(title="Daily Hours", gridcolor=c["grid"], zerolinecolor=c["grid"],
                   tickfont=dict(color=c["subtext"])),
        yaxis2=dict(
            title="Cumulative Hours",
            overlaying="y",
            side="right",
            tickfont=dict(color=CVU_PALETTE[0]),
            showgrid=False,
            zerolinecolor=c["grid"],
        ),
        legend=dict(orientation="h", y=1.08, font=dict(color=c["text"])),
    ))
    st.plotly_chart(fig_line, use_container_width=True)

    # Detailed log (collapsed)
    with st.expander("View detailed entry log"):
        tbl = (
            pdf.assign(Date=pdf["date"].dt.date)
               .groupby(["Date", "category", "task"])["hours"]
               .sum()
               .round(2)
               .reset_index()
               .sort_values(["Date", "category"])
               .rename(columns={"category": "Category", "task": "Task", "hours": "Hours"})
        )
        tbl["Task"] = tbl["Task"].apply(short_name)
        st.dataframe(tbl, hide_index=True, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: TEAM OVERVIEW (manager dashboard)
# ══════════════════════════════════════════════════════════════════════════════

def view_team(df: pd.DataFrame):
    st.subheader("Team Overview Dashboard")
    c = _chart_colors()
    if df.empty:
        st.info("No time data found. Make sure the Excel files are in the same folder as the app.")
        return

    # ── Filters ───────────────────────────────────────────────────────────────
    fc1, fc2 = st.columns(2)
    with fc1:
        quarters  = ["All"] + sorted(df["quarter"].unique().tolist())
        sel_q     = st.selectbox("Quarter", quarters)
    with fc2:
        all_staff = sorted(df["person"].unique())
        sel_staff = st.multiselect("Staff members", all_staff, default=all_staff)

    include_future = st.checkbox(
        "Include future entries (e.g. pre-entered vacation)",
        value=False,
    )

    fdf = df.copy()
    if sel_q != "All":
        fdf = fdf[fdf["quarter"] == sel_q]
    if sel_staff:
        fdf = fdf[fdf["person"].isin(sel_staff)]
    if not include_future:
        fdf = fdf[fdf["date"].dt.date <= date.today()]
    fdf = fdf[fdf["task"].str.startswith("- ")]   # child rows only

    if fdf.empty:
        st.info("Nothing matches the current filters.")
        return

    # ── KPI row ───────────────────────────────────────────────────────────────
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total Hours Logged", f"{fdf['hours'].sum():,.1f}")
    k2.metric("Staff Members",      str(fdf["person"].nunique()))
    k3.metric("Unique Task Codes",  str(fdf["task"].nunique()))
    k4.metric("Days with Entries",  str(fdf["date"].dt.date.nunique()))
    st.divider()

    # Assign consistent colors to categories and staff
    cats          = sorted(fdf["category"].unique())
    cat_color_map = {cat: CVU_PALETTE[i % len(CVU_PALETTE)] for i, cat in enumerate(cats)}
    staff_list    = sorted(fdf["person"].unique())
    staff_colors  = {s: CVU_PALETTE[i % len(CVU_PALETTE)] for i, s in enumerate(staff_list)}

    # ─────────────────────────────────────────────────────────────────────────
    # CHART A + B: Summary donut pair — 900s split & funded/unfunded split
    # Absence codes (120-124) are excluded from both charts.
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Time Allocation Summary")
    non_absence = fdf[~fdf["task"].apply(lambda t: task_subcode(t) in ABSENCE_CODES)].copy()

    donut_left, donut_right = st.columns(2)

    # ── Donut A: 900-series vs all other codes ────────────────────────────────
    with donut_left:
        hrs_900   = non_absence[non_absence["code"] == "900"]["hours"].sum()
        hrs_other = non_absence[non_absence["code"] != "900"]["hours"].sum()
        total_ab  = hrs_900 + hrs_other

        if total_ab > 0:
            figA = go.Figure(go.Pie(
                labels=["900-Series", "Other Codes"],
                values=[hrs_900, hrs_other],
                hole=0.58,
                marker_colors=[CVU_GREEN, CVU_PALETTE[0]],
                hovertemplate="%{label}<br>%{value:.1f} hrs (%{percent})<extra></extra>",
                textinfo="percent",
                textfont=dict(color=c["text"], size=13, family="Inter, Arial, sans-serif"),
            ))
            figA.update_layout(**_chart_base(
                height=300,
                title=dict(text="Research (900s) vs Other",
                           font=dict(color=c["text"], size=13), x=0.5, xanchor="center"),
                showlegend=True,
                legend=dict(orientation="h", y=-0.08, x=0.5, xanchor="center"),
                margin=dict(t=50, b=40, l=20, r=20),
            ))
            st.plotly_chart(figA, use_container_width=True)
        else:
            st.info("No data for this chart.")

    # ── Donut B: Funded vs Unfunded ───────────────────────────────────────────
    with donut_right:
        non_absence["_funded"] = non_absence["task"].apply(
            lambda t: task_subcode(t) in FUNDED_CODES
        )
        hrs_funded   = non_absence[non_absence["_funded"]]["hours"].sum()
        hrs_unfunded = non_absence[~non_absence["_funded"]]["hours"].sum()
        total_fu     = hrs_funded + hrs_unfunded

        if total_fu > 0:
            figB = go.Figure(go.Pie(
                labels=["Funded", "Unfunded"],
                values=[hrs_funded, hrs_unfunded],
                hole=0.58,
                marker_colors=[CVU_GREEN, CVU_PALETTE[1]],
                hovertemplate="%{label}<br>%{value:.1f} hrs (%{percent})<extra></extra>",
                textinfo="percent",
                textfont=dict(color=c["text"], size=13, family="Inter, Arial, sans-serif"),
            ))
            figB.update_layout(**_chart_base(
                height=300,
                title=dict(text="Funded vs Unfunded Time",
                           font=dict(color=c["text"], size=13), x=0.5, xanchor="center"),
                showlegend=True,
                legend=dict(orientation="h", y=-0.08, x=0.5, xanchor="center"),
                margin=dict(t=50, b=40, l=20, r=20),
            ))
            st.plotly_chart(figB, use_container_width=True)
        else:
            st.info("No data for this chart.")

    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # CHART 1: Hours by category (horizontal bar)
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Hours by Category")
    cat_df = (
        fdf.groupby("category")["hours"]
           .sum()
           .reset_index()
           .sort_values("hours")
    )
    fig1 = go.Figure(go.Bar(
        y=cat_df["category"],
        x=cat_df["hours"],
        orientation="h",
        marker_color=[cat_color_map.get(c, CVU_GREEN) for c in cat_df["category"]],
        hovertemplate="%{y}<br>Hours: %{x:.1f}<extra></extra>",
    ))
    fig1.update_layout(**_chart_base(
        height=max(300, len(cat_df) * 55),
        yaxis=dict(automargin=True, tickfont=dict(color=c["text"], size=11),
                   gridcolor=c["grid"], zerolinecolor=c["grid"],
                   tickmode="array",
                   tickvals=cat_df["category"].tolist(),
                   ticktext=cat_df["category"].tolist()),
        xaxis=dict(title="Total Hours", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
    ))
    st.plotly_chart(fig1, use_container_width=True)
    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # CHART 2: Hours over time — stacked bar by staff member
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Hours Over Time")
    agg  = st.radio("Group by", ["Day", "Week", "Month"], horizontal=True, key="agg")
    pcol = {"Day": "date", "Week": "week", "Month": "month"}[agg]

    time_df = fdf.groupby([pcol, "person"])["hours"].sum().reset_index()
    time_df["label"] = (
        time_df["date"].dt.strftime("%Y-%m-%d")
        if agg == "Day"
        else time_df[pcol].astype(str)
    )

    fig2 = go.Figure()
    for person in staff_list:
        pdata = time_df[time_df["person"] == person]
        fig2.add_trace(go.Bar(
            x=pdata["label"],
            y=pdata["hours"],
            name=person,
            marker_color=staff_colors[person],
            hovertemplate=f"{person}<br>%{{x}}<br>Hours: %{{y:.1f}}<extra></extra>",
        ))
    # Full-team capacity line: 4 staff × 8 hrs — bar above = overtime
    full_day_hrs = 8.0 * len(STAFF)
    fig2.add_hline(
        y=full_day_hrs,
        line_dash="dot",
        line_color=c["refline"],
        opacity=0.5,
        annotation_text=f"{full_day_hrs:.0f} hrs (full day)",
        annotation_position="top right",
        annotation_font_color=c["refline"],
        annotation_font_size=11,
    )
    fig2.update_layout(**_chart_base(
        barmode="stack",
        height=360,
        xaxis=dict(tickangle=-40, tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        yaxis=dict(title="Hours", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
    ))
    st.plotly_chart(fig2, use_container_width=True)
    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # CHART 3: Hours per staff member by category
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Hours per Staff Member by Category")
    pp_df = fdf.groupby(["person", "category"])["hours"].sum().reset_index()

    fig3 = go.Figure()
    for cat in cats:
        cdata = pp_df[pp_df["category"] == cat]
        fig3.add_trace(go.Bar(
            x=cdata["person"],
            y=cdata["hours"],
            name=cat,
            marker_color=cat_color_map[cat],
            hovertemplate=f"{cat}<br>%{{x}}<br>Hours: %{{y:.1f}}<extra></extra>",
        ))
    fig3.update_layout(**_chart_base(
        barmode="stack",
        height=420,
        xaxis=dict(tickfont=dict(color=c["text"], size=12),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        yaxis=dict(title="Hours", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=10)),
    ))
    st.plotly_chart(fig3, use_container_width=True)
    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # CHART 4: Cumulative YTD hours by staff member
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Cumulative Hours YTD (by Staff Member)")
    daily_person = (
        fdf.groupby(["date", "person"])["hours"]
           .sum()
           .reset_index()
           .sort_values("date")
    )
    fig4 = go.Figure()
    for person in staff_list:
        pdata = daily_person[daily_person["person"] == person].copy()
        pdata["cumulative"] = pdata["hours"].cumsum()
        fig4.add_trace(go.Scatter(
            x=pdata["date"],
            y=pdata["cumulative"],
            name=person,
            mode="lines",
            line=dict(color=staff_colors[person], width=2),
            hovertemplate=f"{person}<br>%{{x|%b %d}}<br>Total: %{{y:.1f}} hrs<extra></extra>",
        ))
    fig4.update_layout(**_chart_base(
        height=340,
        xaxis=dict(title="", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        yaxis=dict(title="Cumulative Hours", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        legend=dict(orientation="h", y=1.08),
    ))
    st.plotly_chart(fig4, use_container_width=True)
    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # CHART 5: Weekly team pace
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Weekly Team Pace")
    weekly = fdf.groupby("week")["hours"].sum().reset_index().sort_values("week")

    # Average only weeks up to and including the current week
    current_week_str = str(pd.Period(date.today(), freq="W"))
    completed_weeks  = weekly[weekly["week"] <= current_week_str]
    avg_weekly       = completed_weeks["hours"].mean() if not completed_weeks.empty else 0

    fig5 = go.Figure()
    fig5.add_trace(go.Bar(
        x=weekly["week"],
        y=weekly["hours"],
        marker_color=CVU_GREEN,
        opacity=0.85,
        name="Weekly Hours",
        hovertemplate="Week: %{x}<br>Hours: %{y:.1f}<extra></extra>",
    ))
    # Draw the average line only across completed weeks (not future empty weeks)
    if not completed_weeks.empty:
        fig5.add_trace(go.Scatter(
            x=[completed_weeks["week"].iloc[0], completed_weeks["week"].iloc[-1]],
            y=[avg_weekly, avg_weekly],
            mode="lines",
            line=dict(color=CVU_PALETTE[0], width=2, dash="dot"),
            name=f"Avg {avg_weekly:.1f} hrs / wk",
            hovertemplate=f"Avg (completed weeks): {avg_weekly:.1f} hrs<extra></extra>",
        ))
    fig5.update_layout(**_chart_base(
        height=320,
        xaxis=dict(tickangle=-40, tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        yaxis=dict(title="Hours", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
    ))
    st.plotly_chart(fig5, use_container_width=True)
    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # CHART 6: Estimated cost by category
    # Note: uses a fixed blended rate. Update BLENDED_RATE below as needed.
    # ─────────────────────────────────────────────────────────────────────────
    st.markdown("##### Estimated Cost by Category")

    BLENDED_RATE = 91.03   # $/hr — update this to reflect your actual blended rate

    cost_df = fdf.groupby("category")["hours"].sum().reset_index()
    cost_df["est_cost"] = cost_df["hours"] * BLENDED_RATE
    cost_df = cost_df.sort_values("est_cost")

    fig6 = go.Figure(go.Bar(
        y=cost_df["category"],
        x=cost_df["est_cost"],
        orientation="h",
        marker_color=[cat_color_map.get(c, CVU_GREEN) for c in cost_df["category"]],
        hovertemplate="%{y}<br>Est. Cost: $%{x:,.0f}<extra></extra>",
    ))
    fig6.update_layout(**_chart_base(
        height=max(300, len(cost_df) * 55),
        yaxis=dict(automargin=True, tickfont=dict(color=c["text"], size=11),
                   gridcolor=c["grid"], zerolinecolor=c["grid"],
                   tickmode="array",
                   tickvals=cost_df["category"].tolist(),
                   ticktext=cost_df["category"].tolist()),
        xaxis=dict(title=f"Estimated Cost (USD @ ${BLENDED_RATE:.0f}/hr)",
                   tickprefix="$", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
    ))
    st.plotly_chart(fig6, use_container_width=True)
    st.divider()

    # ── Raw data table (collapsed) ────────────────────────────────────────────
    with st.expander("View detailed data table"):
        tbl = (
            fdf.groupby(["person", "category", "task"])["hours"]
               .sum()
               .round(2)
               .reset_index()
               .sort_values(["person", "category"])
        )
        tbl["task"] = tbl["task"].apply(short_name)
        st.dataframe(tbl, hide_index=True, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: FINANCIAL KPIs (admin only)
# ══════════════════════════════════════════════════════════════════════════════

def view_financial_kpis():
    c = _chart_colors()
    txns, settings, all_codes = load_finances()

    # ── Settings & Goals ──────────────────────────────────────────────────────
    with st.expander("Settings & Goals", expanded=txns.empty):
        s1, s2 = st.columns(2)
        ann_staff = s1.number_input(
            "Annual Total Staff Cost ($)",
            min_value=0.0, step=1000.0,
            value=float(settings.get("annual_staff_cost", 0.0)),
            format="%.2f",
            key="fin_staff_cost",
        )
        ann_goal = s2.number_input(
            "Annual Income Goal ($)",
            min_value=0.0, step=1000.0,
            value=float(settings.get("annual_income_goal", 0.0)),
            format="%.2f",
            key="fin_income_goal",
        )

        st.markdown(
            "<p style='color:#9E9E9E;font-size:0.85rem;margin-top:12px'>"
            "Per-Code Income Goals (optional — leave at 0 to skip)</p>",
            unsafe_allow_html=True,
        )
        code_goals  = {}
        code_items  = list(all_codes.items())
        goal_cols   = st.columns(3)
        for i, (code, name) in enumerate(code_items):
            with goal_cols[i % 3]:
                goal_key = f"goal_{code}"
                code_goals[code] = st.number_input(
                    f"{code} — {name[:28]}",
                    min_value=0.0, step=500.0,
                    value=float(settings.get(goal_key, 0.0)),
                    format="%.2f",
                    key=f"fin_goal_{code}",
                    label_visibility="visible",
                )

        if st.button("Save Settings", type="primary"):
            new_settings = {
                "annual_staff_cost":  ann_staff,
                "annual_income_goal": ann_goal,
            }
            for code, goal in code_goals.items():
                if goal > 0:
                    new_settings[f"goal_{code}"] = goal
            custom_codes = {k: v for k, v in all_codes.items() if k not in ACCOUNTING_CODES}
            ok, msg = save_finances(txns, new_settings, custom_codes)
            if ok:
                st.toast("Settings saved.")
                st.rerun()
            else:
                st.error(msg)

    # ── Burn-rate KPI metrics ─────────────────────────────────────────────────
    today         = date.today()
    days_lapsed   = (today - date(today.year, 1, 1)).days + 1
    ann_staff_val = float(settings.get("annual_staff_cost", 0.0))
    daily_burn    = ann_staff_val / 365 if ann_staff_val > 0 else 0.0
    ytd_staff     = daily_burn * days_lapsed

    income_txns  = txns[txns["type"] == "Income"]  if not txns.empty else pd.DataFrame(columns=TXNS_COLS)
    expense_txns = txns[txns["type"] == "Expense"] if not txns.empty else pd.DataFrame(columns=TXNS_COLS)

    paid_income  = income_txns[income_txns["status"] == "Paid"]["amount"].sum()         if not income_txns.empty else 0.0
    soft_income  = income_txns[~income_txns["status"].isin(["Paid"])]["amount"].sum()   if not income_txns.empty else 0.0
    total_exp    = expense_txns["amount"].sum()                                          if not expense_txns.empty else 0.0
    net          = paid_income - total_exp - ytd_staff

    st.markdown("##### Financial Overview")
    k1, k2, k3 = st.columns(3)
    k1.metric("Daily Burn Rate",    f"${daily_burn:,.0f} / day" if daily_burn else "Set staff cost above")
    k2.metric("YTD Staff Cost",     f"${ytd_staff:,.0f}")
    k3.metric("Income — Paid",      f"${paid_income:,.0f}")

    k4, k5, k6 = st.columns(3)
    k4.metric("Income — Pipeline",  f"${soft_income:,.0f}")
    k5.metric("Other Expenses YTD", f"${total_exp:,.0f}")
    k6.metric("Net (Paid - Costs)", f"${net:,.0f}")
    st.divider()

    # ── Add Transaction ───────────────────────────────────────────────────────
    with st.expander("Add Transaction", expanded=False):
        custom_codes = {k: v for k, v in all_codes.items() if k not in ACCOUNTING_CODES}

        ac1, ac2, ac3 = st.columns([1, 1, 1])
        txn_date = ac1.date_input("Date", value=today, key="txn_date")
        txn_type = ac2.selectbox("Type", ["Income", "Expense"], key="txn_type")
        txn_amt  = ac3.number_input("Amount ($)", min_value=0.0, step=100.0,
                                    format="%.2f", key="txn_amt")

        code_options     = [f"{k} — {v}" for k, v in all_codes.items()] + ["+ Add new code"]
        bc1, bc2         = st.columns(2)
        sel_code_str     = bc1.selectbox("Accounting Code", code_options, key="txn_code_sel")
        txn_desc         = bc2.text_input("Description", key="txn_desc")

        if sel_code_str == "+ Add new code":
            nc1, nc2      = st.columns(2)
            new_code_id   = nc1.text_input("Code (e.g. 918)", key="new_code_id")
            new_code_name = nc2.text_input("Code name", key="new_code_name")
            txn_code      = new_code_id.strip()
            txn_code_name = new_code_name.strip()
            save_new_code = st.checkbox("Save this code for future use", value=True, key="save_new_code")
        else:
            parts         = sel_code_str.split(" — ", 1)
            txn_code      = parts[0].strip()
            txn_code_name = parts[1].strip() if len(parts) > 1 else txn_code
            save_new_code = False

        txn_status = ""
        if txn_type == "Income":
            txn_status = st.selectbox("Status", INCOME_STATUSES, key="txn_status")

        txn_notes = st.text_input("Notes (optional)", key="txn_notes")

        if st.button("Add Transaction", type="primary", use_container_width=True):
            if txn_amt <= 0:
                st.error("Please enter an amount greater than 0.")
            elif not txn_code:
                st.error("Please select or enter an accounting code.")
            else:
                new_id  = int(txns["id"].max() + 1) if not txns.empty and txns["id"].notna().any() else 1
                new_row = pd.DataFrame([{
                    "id":          new_id,
                    "date":        pd.Timestamp(txn_date),
                    "type":        txn_type,
                    "amount":      txn_amt,
                    "code":        txn_code,
                    "code_name":   txn_code_name,
                    "description": txn_desc,
                    "status":      txn_status,
                    "notes":       txn_notes,
                }])
                updated_txns   = pd.concat([txns, new_row], ignore_index=True)
                updated_custom = dict(custom_codes)
                if save_new_code and txn_code and txn_code not in ACCOUNTING_CODES:
                    updated_custom[txn_code] = txn_code_name
                ok, msg = save_finances(updated_txns, dict(settings), updated_custom)
                if ok:
                    st.toast("Transaction added.")
                    st.rerun()
                else:
                    st.error(msg)

    if txns.empty:
        st.info("No transactions yet — use the form above to add income or expenses.")
        return

    st.divider()

    # ── Chart row 1: Income by status donut  +  Progress to annual goal ───────
    st.markdown("##### Income Pipeline")
    ch1, ch2 = st.columns(2)

    with ch1:
        if not income_txns.empty:
            status_grp = income_txns.groupby("status")["amount"].sum().reset_index()
            figA = go.Figure(go.Pie(
                labels=status_grp["status"],
                values=status_grp["amount"],
                hole=0.58,
                marker_colors=[STATUS_COLORS.get(s, CVU_GRAY) for s in status_grp["status"]],
                hovertemplate="%{label}<br>$%{value:,.0f} (%{percent})<extra></extra>",
                textinfo="percent",
                textfont=dict(color=c["text"], size=12, family="Inter, Arial, sans-serif"),
            ))
            figA.update_layout(**_chart_base(
                height=300,
                title=dict(text="Income by Status",
                           font=dict(color=c["text"], size=13), x=0.5, xanchor="center"),
                showlegend=True,
                legend=dict(orientation="h", y=-0.12, x=0.5, xanchor="center"),
                margin=dict(t=50, b=55, l=20, r=20),
            ))
            st.plotly_chart(figA, use_container_width=True)
        else:
            st.info("No income transactions yet.")

    with ch2:
        annual_goal_val = float(settings.get("annual_income_goal", 0.0))
        if annual_goal_val > 0 and not income_txns.empty:
            seg_data = {s: income_txns[income_txns["status"] == s]["amount"].sum()
                        for s in INCOME_STATUSES}

            total_all       = sum(seg_data.values())
            total_confirmed = sum(seg_data.get(s, 0)
                                  for s in ("Paid", "Invoiced", "Contracted"))
            pct_all       = min(100.0, total_all       / annual_goal_val * 100)
            pct_confirmed = min(100.0, total_confirmed / annual_goal_val * 100)

            figB = go.Figure()
            for s in INCOME_STATUSES:
                figB.add_trace(go.Bar(
                    x=[seg_data.get(s, 0)],
                    y=["Income"],
                    orientation="h",
                    name=s,
                    marker_color=STATUS_COLORS.get(s, CVU_GRAY),
                    hovertemplate=f"{s}: $%{{x:,.0f}}<extra></extra>",
                ))
            figB.add_vline(
                x=annual_goal_val,
                line_dash="dot",
                line_color=c["refline"],
                annotation_text=f"Goal  ${annual_goal_val:,.0f}",
                annotation_font_color=c["refline"],
                annotation_position="top right",
            )
            figB.update_layout(**_chart_base(
                barmode="stack",
                height=220,
                title=dict(text="Progress to Annual Income Goal",
                           font=dict(color=c["text"], size=13), x=0.5, xanchor="center"),
                showlegend=True,
                legend=dict(orientation="h", y=-0.25, x=0.5, xanchor="center"),
                margin=dict(t=50, b=90, l=20, r=20),
                xaxis=dict(tickprefix="$", tickfont=dict(color=c["subtext"]),
                           gridcolor=c["grid"], zerolinecolor=c["grid"]),
                yaxis=dict(showticklabels=False, gridcolor=c["grid"]),
            ))
            st.plotly_chart(figB, use_container_width=True)

            # Percentage labels below the chart
            pa, pb = st.columns(2)
            pa.metric("All income vs goal",
                      f"{pct_all:.1f}%",
                      help="Every status (Paid + Invoiced + Contracted + Verbal + Pipeline) as a % of the annual goal")
            pb.metric("Confirmed vs goal",
                      f"{pct_confirmed:.1f}%",
                      help="Only Paid, Invoiced, and Contracted — excludes Verbal and Pipeline")

        elif annual_goal_val == 0:
            st.info("Set an annual income goal in Settings to see progress here.")

    st.divider()

    # ── Chart 2: Income vs Expenses by code ───────────────────────────────────
    st.markdown("##### Income vs Expenses by Code")
    inc_by_code = (income_txns.groupby("code")["amount"].sum().reset_index()
                   .rename(columns={"amount": "income"})
                   if not income_txns.empty else pd.DataFrame(columns=["code", "income"]))
    exp_by_code = (expense_txns.groupby("code")["amount"].sum().reset_index()
                   .rename(columns={"amount": "expenses"})
                   if not expense_txns.empty else pd.DataFrame(columns=["code", "expenses"]))

    by_code = pd.merge(inc_by_code, exp_by_code, on="code", how="outer").fillna(0)
    if not by_code.empty:
        by_code["label"] = by_code["code"].map(all_codes).fillna(by_code["code"])
        by_code["label"] = by_code.apply(lambda r: f"{r['code']} — {r['label']}", axis=1)
        by_code = by_code.sort_values("income", ascending=False)

        figC = go.Figure()
        figC.add_trace(go.Bar(
            x=by_code["label"], y=by_code["income"],
            name="Income", marker_color=CVU_GREEN,
            hovertemplate="%{x}<br>Income: $%{y:,.0f}<extra></extra>",
        ))
        figC.add_trace(go.Bar(
            x=by_code["label"], y=by_code["expenses"],
            name="Expenses", marker_color=CVU_PALETTE[5],
            hovertemplate="%{x}<br>Expenses: $%{y:,.0f}<extra></extra>",
        ))
        figC.update_layout(**_chart_base(
            barmode="group",
            height=360,
            xaxis=dict(tickangle=-35, tickfont=dict(color=c["subtext"], size=10),
                       gridcolor=c["grid"], zerolinecolor=c["grid"]),
            yaxis=dict(title="Amount ($)", tickprefix="$", tickfont=dict(color=c["subtext"]),
                       gridcolor=c["grid"], zerolinecolor=c["grid"]),
            legend=dict(orientation="h", y=1.08),
        ))
        st.plotly_chart(figC, use_container_width=True)

    # Per-code goal progress (only show codes that have a goal set)
    code_goal_rows = [(c, float(settings[f"goal_{c}"])) for c in all_codes
                      if f"goal_{c}" in settings and float(settings[f"goal_{c}"]) > 0]
    if code_goal_rows:
        st.markdown("##### Per-Code Income Progress vs Goal")
        goal_labels, goal_vals, actual_vals = [], [], []
        for code, goal in sorted(code_goal_rows):
            actual = income_txns[income_txns["code"] == code]["amount"].sum() if not income_txns.empty else 0.0
            goal_labels.append(f"{code} — {all_codes.get(code, code)}")
            goal_vals.append(goal)
            actual_vals.append(actual)

        figD = go.Figure()
        # One bar segment per income status, coloured to match the pipeline donut
        for status in INCOME_STATUSES:
            status_vals = []
            for code, _ in sorted(code_goal_rows):
                v = (income_txns[(income_txns["code"] == code) &
                                  (income_txns["status"] == status)]["amount"].sum()
                     if not income_txns.empty else 0.0)
                status_vals.append(v)
            figD.add_trace(go.Bar(
                y=goal_labels, x=status_vals,
                orientation="h",
                name=status,
                marker_color=STATUS_COLORS[status],
                hovertemplate=f"{status}: $%{{x:,.0f}}<extra></extra>",
            ))

        # Remaining-to-goal segment (muted)
        figD.add_trace(go.Bar(
            y=goal_labels,
            x=[max(0.0, g - a) for g, a in zip(goal_vals, actual_vals)],
            orientation="h",
            name="Remaining",
            marker_color=CVU_SURFACE,
            marker_line=dict(color=CVU_BORDER, width=1),
            hovertemplate="Remaining to goal: $%{x:,.0f}<extra></extra>",
        ))

        # Percentage-complete annotation — only Paid, Invoiced, Contracted count
        CONFIRMED_STATUSES = {"Paid", "Invoiced", "Contracted"}
        for (code, _), label, goal in zip(sorted(code_goal_rows), goal_labels, goal_vals):
            confirmed = (
                income_txns[
                    (income_txns["code"] == code) &
                    (income_txns["status"].isin(CONFIRMED_STATUSES))
                ]["amount"].sum()
                if not income_txns.empty else 0.0
            )
            pct = min(100.0, confirmed / goal * 100) if goal > 0 else 0.0
            figD.add_annotation(
                x=goal, y=label,
                text=f"  {pct:.0f}%",
                showarrow=False,
                xanchor="left",
                font=dict(color=c["text"], size=11, family="Inter, Arial, sans-serif"),
            )

        figD.update_layout(**_chart_base(
            barmode="stack",
            height=max(300, len(code_goal_rows) * 54),
            yaxis=dict(automargin=True, tickfont=dict(color=c["text"], size=11),
                       gridcolor=c["grid"], zerolinecolor=c["grid"]),
            xaxis=dict(title="Amount ($)", tickprefix="$", tickfont=dict(color=c["subtext"]),
                       gridcolor=c["grid"], zerolinecolor=c["grid"]),
            legend=dict(orientation="h", y=1.06),
        ))
        st.plotly_chart(figD, use_container_width=True)

    st.divider()

    # ── Chart 3: Monthly cash flow (income, staff burn, other expenses, net) ──
    st.markdown("##### Monthly Cash Flow")

    # Build every month from Jan through today so staff burn always appears
    def _month_staff_cost(month_str: str) -> float:
        """Daily burn × days worked in the month (full month if past, days-to-date if current)."""
        y, m = int(month_str[:4]), int(month_str[5:7])
        if date(y, m, 1) > today:
            return 0.0
        if y == today.year and m == today.month:
            days = today.day
        else:
            days = calendar.monthrange(y, m)[1]
        return daily_burn * days

    year_months = [f"{today.year}-{m:02d}" for m in range(1, today.month + 1)]

    txns_m = txns.copy()
    txns_m["month"] = txns_m["date"].dt.to_period("M").astype(str)
    monthly_inc = (txns_m[txns_m["type"] == "Income"]
                   .groupby("month")["amount"].sum())
    monthly_exp = (txns_m[txns_m["type"] == "Expense"]
                   .groupby("month")["amount"].sum())

    # Merge transaction months with the full year-to-date month list
    all_months = sorted(set(year_months)
                        | set(monthly_inc.index)
                        | set(monthly_exp.index))
    # Keep only current year
    all_months = [m for m in all_months if m.startswith(str(today.year))]

    monthly_staff_costs = [_month_staff_cost(m) for m in all_months]
    net_per_month = [
        monthly_inc.get(m, 0) - monthly_exp.get(m, 0) - _month_staff_cost(m)
        for m in all_months
    ]

    figE = go.Figure()
    figE.add_trace(go.Bar(
        x=all_months,
        y=[monthly_inc.get(m, 0) for m in all_months],
        name="Income",
        marker_color=CVU_GREEN,
        hovertemplate="<b>%{x}</b><br>Income: $%{y:,.0f}<extra></extra>",
    ))
    figE.add_trace(go.Bar(
        x=all_months,
        y=[-c for c in monthly_staff_costs],
        name="Staff Cost",
        marker_color=CVU_PALETTE[1],
        hovertemplate="<b>%{x}</b><br>Staff Cost: $%{customdata:,.0f}<extra></extra>",
        customdata=monthly_staff_costs,
    ))
    figE.add_trace(go.Bar(
        x=all_months,
        y=[-monthly_exp.get(m, 0) for m in all_months],
        name="Other Expenses",
        marker_color=CVU_PALETTE[5],
        hovertemplate="<b>%{x}</b><br>Other Expenses: $%{customdata:,.0f}<extra></extra>",
        customdata=[monthly_exp.get(m, 0) for m in all_months],
    ))
    figE.add_trace(go.Scatter(
        x=all_months,
        y=net_per_month,
        name="Net",
        mode="lines+markers+text",
        line=dict(color=CVU_PALETTE[0], width=2),
        marker=dict(size=7, color=CVU_PALETTE[0]),
        text=[f"${v:,.0f}" for v in net_per_month],
        textposition="top center",
        textfont=dict(color=CVU_PALETTE[0], size=10),
        hovertemplate="<b>%{x}</b><br>Net: $%{y:,.0f}<extra></extra>",
    ))
    figE.update_layout(**_chart_base(
        barmode="relative",
        height=360,
        xaxis=dict(tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        yaxis=dict(title="Amount ($)", tickprefix="$", tickfont=dict(color=c["subtext"]),
                   gridcolor=c["grid"], zerolinecolor=c["grid"]),
        legend=dict(orientation="h", y=1.08),
    ))
    st.plotly_chart(figE, use_container_width=True)

    st.divider()

    # ── Transaction Table (editable) ──────────────────────────────────────────
    st.markdown("##### All Transactions")
    st.caption("Edit any cell directly, then click Save Changes below.")

    tf1, tf2, tf3 = st.columns(3)
    f_type   = tf1.selectbox("Filter: Type",   ["All", "Income", "Expense"], key="tbl_type")
    f_status = tf2.selectbox("Filter: Status", ["All"] + INCOME_STATUSES,   key="tbl_status")
    f_code   = tf3.selectbox("Filter: Code",   ["All"] + sorted(txns["code"].unique().tolist()), key="tbl_code")

    # Build editable subset — keep original txns index so we can merge back
    mask = pd.Series(True, index=txns.index)
    if f_type   != "All": mask &= txns["type"]   == f_type
    if f_status != "All": mask &= txns["status"] == f_status
    if f_code   != "All": mask &= txns["code"]   == f_code

    edit_df = txns[mask].sort_values("date", ascending=False).drop(columns=["id"]).copy()
    edit_df["date"] = edit_df["date"].dt.date   # DateColumn expects date objects

    code_options   = list(all_codes.keys())
    status_options = [""] + INCOME_STATUSES

    edited = st.data_editor(
        edit_df,
        use_container_width=True,
        num_rows="fixed",
        hide_index=True,
        column_config={
            "date":        st.column_config.DateColumn("Date", required=True),
            "type":        st.column_config.SelectboxColumn(
                               "Type", options=["Income", "Expense"], required=True),
            "amount":      st.column_config.NumberColumn(
                               "Amount ($)", min_value=0.0, step=0.01, format="$%.2f"),
            "code":        st.column_config.SelectboxColumn("Code", options=code_options),
            "code_name":   st.column_config.TextColumn("Code Name"),
            "description": st.column_config.TextColumn("Description"),
            "status":      st.column_config.SelectboxColumn("Status", options=status_options),
            "notes":       st.column_config.TextColumn("Notes"),
        },
    )

    if st.button("Save Changes", type="primary", use_container_width=True):
        # Coerce types and auto-fill code_name when code changes
        edited["date"]   = pd.to_datetime(edited["date"])
        edited["amount"] = pd.to_numeric(edited["amount"], errors="coerce").fillna(0.0)
        edited["status"] = edited["status"].fillna("")
        edited["notes"]  = edited["notes"].fillna("")
        edited["code_name"] = edited.apply(
            lambda r: all_codes.get(r["code"], r.get("code_name", r["code"])),
            axis=1,
        )

        # Merge edited rows back into the full dataset using the preserved index
        updated_txns = txns.copy()
        updated_txns.loc[edited.index, edit_df.columns] = edited
        updated_txns["id"] = range(1, len(updated_txns) + 1)

        custom_codes = {k: v for k, v in all_codes.items() if k not in ACCOUNTING_CODES}
        ok, msg = save_finances(updated_txns, dict(settings), custom_codes)
        if ok:
            st.toast("Changes saved.")
            st.rerun()
        else:
            st.error(msg)


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: CONTENT KPIs (admin only)
# ══════════════════════════════════════════════════════════════════════════════

def _content_progress(row):
    """
    Returns (total_pct, p1_fill_pct, p2_fill_pct, p3_fill_pct).
    Phase weights: Draft=30%, Layout=60%, Production=10%.
    Fill pcts are 0–100 within each phase segment (for the progress bar).
    """
    fmt = str(row.get("format", "TBD"))

    def filled(key):
        """A milestone counts only if a date is recorded AND that date has passed."""
        v = row.get(key, "")
        if not v or str(v) in ("", "None", "NaT", "nan"):
            return False
        try:
            return date.fromisoformat(str(v)[:10]) <= date.today()
        except (ValueError, TypeError):
            return False

    # Phase 1 — Draft (30% of total)
    p1_done  = sum(filled(k) for k in ("draft_delivered", "draft_commented", "draft_completed"))
    p1_fill  = round((p1_done / 3) * 100)
    p1_total = (p1_done / 3) * 30

    # Phase 2 — Layout (60% of total)
    p2_done  = sum(filled(k) for k in (
        "layout1_delivered", "layout1_commented",
        "layout2_delivered", "layout2_approved",
    ))
    p2_fill  = round((p2_done / 4) * 100)
    p2_total = (p2_done / 4) * 60

    # Phase 3 — Production (10% of total)
    # Digital-only: only go-live counts (10%)
    # Print-only:   only print date counts (10%)
    # Both/TBD:     print date = 5%, go-live = 5%
    if fmt == "Digital":
        go       = filled("go_live_date")
        p3_fill  = 100 if go else 0
        p3_total = 10.0 if go else 0.0
    elif fmt == "Print":
        pr       = filled("print_date")
        p3_fill  = 100 if pr else 0
        p3_total = 10.0 if pr else 0.0
    else:
        pr       = filled("print_date")
        go       = filled("go_live_date")
        p3_fill  = round(((pr + go) / 2) * 100)
        p3_total = (pr + go) * 5.0

    total = round(p1_total + p2_total + p3_total)
    return total, p1_fill, p2_fill, p3_fill


def _progress_from_override(pct: int):
    """
    Map a manual 0–100 override to (total, p1_fill, p2_fill, p3_fill).
    Fills the three phase segments sequentially (Draft→Layout→Prod)
    so the bar reads left-to-right like real progress.
    """
    pct = max(0, min(100, int(pct)))
    if pct <= 30:
        p1 = round(pct / 30 * 100)
        p2, p3 = 0, 0
    elif pct <= 90:
        p1 = 100
        p2 = round((pct - 30) / 60 * 100)
        p3 = 0
    else:
        p1, p2 = 100, 100
        p3 = round((pct - 90) / 10 * 100)
    return pct, p1, p2, p3


def _compute_project_pct(row) -> int:
    """Return the effective % complete for a project row (override takes priority)."""
    try:
        ov = int(float(row.get("pct_override", 0) or 0))
    except (ValueError, TypeError):
        ov = 0
    if ov > 0:
        return ov
    total, _, _, _ = _content_progress(row)
    return total


def _next_milestone(proj: dict):
    """
    Return (date_str, display_name) for the earliest future milestone,
    or None if no upcoming dates are set.
    Respects format: print_date excluded for Digital, go_live_date excluded for Print.
    """
    fmt = str(proj.get("format", "TBD"))
    candidates = [
        ("draft_delivered",   "Draft delivered"),
        ("draft_commented",   "Draft commented"),
        ("draft_completed",   "Draft completed"),
        ("layout1_delivered", "Layout 1 delivered"),
        ("layout1_commented", "Layout 1 commented"),
        ("layout2_delivered", "Layout 2 delivered"),
        ("layout2_approved",  "Layout 2 approved"),
    ]
    if fmt != "Digital":
        candidates.append(("print_date",  "Sent to printer"))
    if fmt != "Print":
        candidates.append(("go_live_date", "Go-live"))

    today = date.today()
    upcoming = []
    for key, label in candidates:
        v = proj.get(key, "")
        if not v or str(v) in ("", "None", "NaT", "nan"):
            continue
        try:
            d = date.fromisoformat(str(v)[:10])
            if d > today:
                upcoming.append((d, label))
        except (ValueError, TypeError):
            continue

    if not upcoming:
        return None
    upcoming.sort(key=lambda x: x[0])
    d, label = upcoming[0]
    return d.isoformat(), label


def _pillar_color(pillar: str) -> str:
    return PILLAR_COLORS.get(pillar, DEFAULT_PILLAR_COLOR)


def _project_card_html(proj: dict, pillar_color: str) -> str:
    """Render one project as a dark card with a pillar-colored progress bar."""
    # Use manual override if set, otherwise auto-calculate from milestone dates
    override_raw = proj.get("pct_override", 0)
    try:
        override = int(float(override_raw)) if override_raw else 0
    except (ValueError, TypeError):
        override = 0

    if override > 0:
        total, p1, p2, p3 = _progress_from_override(override)
        manual_label = (
            "<span style='font-size:0.65rem;color:#9E9E9E;font-weight:400'>"
            " (manual)</span>"
        )
    else:
        total, p1, p2, p3 = _content_progress(proj)
        manual_label = ""

    title      = proj.get("title",            "Untitled")
    status     = proj.get("status",           "")
    conf_pend  = proj.get("confirmed_pending","Confirmed") or "Confirmed"
    pillar     = proj.get("pillar",           "")
    fmt        = proj.get("format",           "")
    owner      = proj.get("owner",            "")
    ptype      = proj.get("type",             "")
    acct       = proj.get("acct_code",        "")
    budget     = proj.get("budget",    0.0) or 0.0
    hours      = proj.get("est_hours", 0.0) or 0.0

    budget_str  = f"${float(budget):,.0f}" if budget else "—"
    hours_str   = f"{float(hours):,.0f} hrs" if hours else "—"
    is_pending  = (conf_pend == "Pending")
    italic_style = "font-style:italic;" if is_pending else ""
    status_line  = f"{conf_pend}, {status}" if status else conf_pend

    next_ms = _next_milestone(proj)
    if next_ms:
        next_date, next_name = next_ms
        next_html = (
            f"<span style='font-size:0.72rem;font-weight:400;"
            f"color:#9E9E9E;margin-left:10px;font-style:normal'>"
            f"Next: {next_date}, {next_name}</span>"
        )
    else:
        next_html = ""

    return f"""
<div style="background:#282828;border-radius:8px;padding:18px 20px 14px;
            border-left:4px solid {pillar_color};margin-bottom:4px;{italic_style}">

  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:2px">
    <span style="font-size:1.05rem;font-weight:700;color:#FCFCFC;
                 line-height:1.3;flex:1;margin-right:8px">{title}</span>
    <span style="background:{pillar_color}22;color:{pillar_color};
                 border:1px solid {pillar_color}55;border-radius:10px;
                 font-size:0.68rem;font-weight:600;padding:2px 8px;
                 white-space:nowrap;flex-shrink:0;font-style:normal">{pillar}</span>
  </div>

  <div style="font-size:0.82rem;color:#C5C5C5;margin-bottom:6px">{status_line}</div>
  <div style="font-size:0.9rem;font-weight:700;color:#B4E817;margin-bottom:10px">
    {total}% complete{manual_label}{next_html}
  </div>

  <!-- Progress bar: 3 phase segments (widths 30% / 60% / 10%) -->
  <div style="display:flex;gap:3px;height:9px;margin-bottom:4px">
    <div style="flex:30;background:#3A3A3A;border-radius:2px;overflow:hidden">
      <div style="height:100%;width:{p1}%;background:{pillar_color};transition:width .3s"></div>
    </div>
    <div style="flex:60;background:#3A3A3A;border-radius:2px;overflow:hidden">
      <div style="height:100%;width:{p2}%;background:{pillar_color};transition:width .3s"></div>
    </div>
    <div style="flex:10;background:#3A3A3A;border-radius:2px;overflow:hidden">
      <div style="height:100%;width:{p3}%;background:{pillar_color};transition:width .3s"></div>
    </div>
  </div>
  <div style="display:flex;gap:3px;margin-bottom:12px">
    <div style="flex:30;font-size:0.65rem;color:#9E9E9E">Draft</div>
    <div style="flex:60;font-size:0.65rem;color:#9E9E9E">Layout</div>
    <div style="flex:10;font-size:0.65rem;color:#9E9E9E;text-align:right">Prod</div>
  </div>

  <!-- Meta row -->
  <div style="display:flex;flex-wrap:wrap;gap:12px;font-size:0.78rem;color:#9E9E9E">
    <span><span style="color:#FCFCFC;font-weight:600">{ptype}</span></span>
    <span>Owner: <span style="color:#FCFCFC;font-weight:600">{owner}</span></span>
    <span>Format: <span style="color:#FCFCFC;font-weight:600">{fmt}</span></span>
    <span>Code: <span style="color:#FCFCFC;font-weight:600">{acct}</span></span>
    <span>Budget: <span style="color:#FCFCFC;font-weight:600">{budget_str}</span></span>
    <span>Est. RTL Hours: <span style="color:#FCFCFC;font-weight:600">{hours_str}</span></span>
  </div>
</div>
"""


def _idx(lst: list, val, default: int = 0) -> int:
    """Return index of val in lst, or default if not found."""
    try:
        return lst.index(str(val)) if val else default
    except ValueError:
        return default


def view_content_kpis():
    c_colors = _chart_colors()
    projects, settings, all_pillars, all_codes = load_content_projects()

    # Hourly rate: auto-derive from Financial KPIs annual staff cost if available.
    # Cache in session_state so load_finances() only fires once per session,
    # not on every widget interaction (which causes page jumps).
    if "cp_auto_rate" not in st.session_state:
        try:
            _, fin_settings, _ = load_finances()
            ann_cost = float(fin_settings.get("annual_staff_cost", 0.0))
            st.session_state["cp_auto_rate"] = (
                round(ann_cost * 0.00012019, 2) if ann_cost > 0 else 0.0
            )
        except Exception:
            st.session_state["cp_auto_rate"] = 0.0
    auto_rate = st.session_state["cp_auto_rate"]

    stored_rate  = float(settings.get("rtl_hourly_rate", 0.0))
    default_rate = stored_rate if stored_rate > 0 else (auto_rate if auto_rate > 0 else 91.03)

    # ── Add / Edit Project ────────────────────────────────────────────────────
    # Determine which project is being edited via a dropdown selector.
    # Embedding the project ID in every widget key guarantees Streamlit
    # creates fresh widgets (honouring value=/index=) whenever the selection
    # changes — no session_state juggling required.
    edit_id   = None
    edit_proj = {}

    if not projects.empty:
        proj_map = {
            f"{row['title']} (#{int(row['id'])})": int(row["id"])
            for _, row in projects.iterrows()
        }
        _ADD_LABEL = "— Add new project —"
        sel_options = [_ADD_LABEL] + list(proj_map.keys())
        sel_label   = st.selectbox(
            "Add a new project, or select one below to edit:",
            sel_options,
            key="cp_edit_sel",
        )
        if sel_label != _ADD_LABEL:
            edit_id = proj_map.get(sel_label)
            if edit_id is not None:
                matches = projects[projects["id"] == edit_id]
                if not matches.empty:
                    edit_proj = matches.iloc[0].to_dict()

    # eid is embedded in every widget key so switching projects gives
    # entirely new keys → Streamlit uses value=/index= fresh each time.
    eid = str(edit_id) if edit_id else "new"

    form_label = (f"Edit — {edit_proj.get('title', '')}" if edit_id
                  else "Add New Project")
    with st.expander(form_label, expanded=(edit_id is not None or projects.empty)):

        # Row 1: Title | Type | Pillar
        f1, f2, f3 = st.columns([3, 2, 2])
        cp_title  = f1.text_input("Title", key=f"cp_title_{eid}",
                                  value=str(edit_proj.get("title", "") or ""))
        cp_type   = f2.selectbox("Type", CONTENT_TYPES, key=f"cp_type_{eid}",
                                 index=_idx(CONTENT_TYPES, edit_proj.get("type")))
        cp_pillar = f3.selectbox("Pillar", all_pillars, key=f"cp_pillar_{eid}",
                                 index=_idx(all_pillars, edit_proj.get("pillar")))

        # Row 2: Accounting Code | Owner | Sponsored
        code_options = [f"{k} — {v}" for k, v in all_codes.items()]
        g1, g2, g3  = st.columns([3, 2, 2])
        cp_code_raw  = g1.selectbox("Accounting Code", code_options,
                                    key=f"cp_code_{eid}",
                                    index=_idx(code_options,
                                               next((f"{k} — {v}" for k, v in all_codes.items()
                                                     if k == edit_proj.get("acct_code")), None)))
        cp_owner     = g2.selectbox("RTL Owner", RTL_OWNERS, key=f"cp_owner_{eid}",
                                    index=_idx(RTL_OWNERS, edit_proj.get("owner")))
        cp_sponsored = g3.selectbox("Sponsored?", ["No", "Yes"],
                                    key=f"cp_sponsored_{eid}",
                                    index=_idx(["No", "Yes"],
                                               edit_proj.get("sponsored", "No")))

        # Row 3 (conditional): Sponsorship Type
        cp_sponsorship_type  = ""
        cp_sponsorship_other = ""
        if cp_sponsored == "Yes":
            sp1, sp2 = st.columns(2)
            cp_sponsorship_type = sp1.selectbox(
                "Sponsorship Type", SPONSORSHIP_TYPES, key=f"cp_spons_type_{eid}",
                index=_idx(SPONSORSHIP_TYPES, edit_proj.get("sponsorship_type")),
            )
            if cp_sponsorship_type == "Other":
                cp_sponsorship_other = sp2.text_input(
                    "Specify sponsorship type", key=f"cp_spons_other_{eid}",
                    value=str(edit_proj.get("sponsorship_other", "") or ""),
                )

        # Row 4: Confirmed/Pending | Status | Format | Content Generator
        h0, h1, h2, h3 = st.columns(4)
        cp_conf_pend = h0.selectbox(
            "Confirmed/Pending",
            ["Confirmed", "Pending"],
            key=f"cp_conf_pend_{eid}",
            index=_idx(["Confirmed", "Pending"],
                       edit_proj.get("confirmed_pending", "Confirmed")),
        )
        cp_status    = h1.selectbox("Priority", PROJECT_STATUSES,
                                    key=f"cp_status_{eid}",
                                    index=_idx(PROJECT_STATUSES, edit_proj.get("status")))
        cp_format    = h2.selectbox("Format", FORMAT_TYPES, key=f"cp_format_{eid}",
                                    index=_idx(FORMAT_TYPES, edit_proj.get("format")))
        cp_generator = h3.selectbox("Content Generator", CONTENT_GENERATORS,
                                    key=f"cp_generator_{eid}",
                                    index=_idx(CONTENT_GENERATORS,
                                               edit_proj.get("content_generator")))

        # Row 5 (conditional): Committee name + Funding Source + Program Partner name
        cp_committee    = ""
        cp_funding      = FUNDING_SOURCES[0]
        cp_partner_name = ""
        fu1, fu2 = st.columns(2)
        if cp_generator == "Committee":
            cp_committee = fu1.text_input(
                "Committee name", key=f"cp_committee_{eid}",
                value=str(edit_proj.get("committee_name", "") or ""),
            )
        cp_funding = fu2.selectbox("Funding Source", FUNDING_SOURCES,
                                   key=f"cp_funding_{eid}",
                                   index=_idx(FUNDING_SOURCES,
                                              edit_proj.get("funding_source")))
        if cp_funding == "Program Partner":
            cp_partner_name = st.text_input(
                "Program partner name", key=f"cp_partner_name_{eid}",
                value=str(edit_proj.get("program_partner_name", "") or ""),
            )

        # Row 6: Budget + Estimated Hours
        b1, b2 = st.columns(2)
        cp_budget = b1.number_input(
            "2026 Budget ($)", min_value=0.0, step=500.0, format="%.2f",
            value=float(edit_proj.get("budget", 0.0) or 0.0),
            key=f"cp_budget_{eid}",
        )
        cp_hours = b2.number_input(
            "Est. RTL Hours", min_value=0.0, step=1.0, format="%.1f",
            value=float(edit_proj.get("est_hours", 0.0) or 0.0),
            key=f"cp_hours_{eid}",
        )

        # Show computed RTL labor cost estimate
        rate = float(settings.get("rtl_hourly_rate", default_rate))
        if cp_hours > 0 and rate > 0:
            labor_cost = cp_hours * rate
            st.caption(f"Estimated RTL labor cost: **${labor_cost:,.0f}** "
                       f"({cp_hours:.0f} hrs × ${rate:.2f}/hr)")

        # ── Schedule ──────────────────────────────────────────────────────────
        st.markdown(
            "<p style='color:#9E9E9E;font-size:0.85rem;margin-top:10px'>"
            "Schedule — enter dates as you reach each milestone "
            "(leave blank until completed)</p>",
            unsafe_allow_html=True,
        )

        def _date_val(key):
            raw = edit_proj.get(key, "")
            if not raw or str(raw) in ("", "None", "NaT", "nan"):
                return None
            try:
                return date.fromisoformat(str(raw)[:10])
            except Exception:
                return None

        st.markdown("**Phase 1 — Draft**")
        d1, d2, d3 = st.columns(3)
        cp_draft_del = d1.date_input("Draft delivered",
                                     value=_date_val("draft_delivered"),
                                     key=f"cp_draft_del_{eid}")
        cp_draft_com = d2.date_input("Draft commented",
                                     value=_date_val("draft_commented"),
                                     key=f"cp_draft_com_{eid}")
        cp_draft_cpd = d3.date_input("Draft completed",
                                     value=_date_val("draft_completed"),
                                     key=f"cp_draft_cpd_{eid}")

        st.markdown("**Phase 2 — Layout**")
        l1, l2, l3, l4 = st.columns(4)
        cp_lay1_del = l1.date_input("Layout 1 delivered",
                                    value=_date_val("layout1_delivered"),
                                    key=f"cp_lay1_del_{eid}")
        cp_lay1_com = l2.date_input("Layout 1 commented",
                                    value=_date_val("layout1_commented"),
                                    key=f"cp_lay1_com_{eid}")
        cp_lay2_del = l3.date_input("Layout 2 delivered",
                                    value=_date_val("layout2_delivered"),
                                    key=f"cp_lay2_del_{eid}")
        cp_lay2_apr = l4.date_input("Layout 2 approved",
                                    value=_date_val("layout2_approved"),
                                    key=f"cp_lay2_apr_{eid}")

        st.markdown("**Phase 3 — Production**")
        p1c, p2c = st.columns(2)
        show_print = cp_format in ("Print", "Both", "TBD")
        show_live  = cp_format in ("Digital", "Both", "TBD")
        cp_print_date = (
            p1c.date_input("Print date", value=_date_val("print_date"),
                           key=f"cp_print_{eid}")
            if show_print else None
        )
        cp_go_live = (
            p2c.date_input("Go-live date", value=_date_val("go_live_date"),
                           key=f"cp_golive_{eid}")
            if show_live else None
        )

        # Notes
        cp_notes = st.text_area("Notes (optional)", key=f"cp_notes_{eid}",
                                value=str(edit_proj.get("notes", "") or ""),
                                height=80)

        # Manual completion override
        st.markdown(
            "<p style='color:#9E9E9E;font-size:0.85rem;margin-top:4px'>"
            "Completion % override — set this if the auto-calculated value "
            "(based on milestone dates) isn't accurate. Set to 0 to go back "
            "to auto-calculate.</p>",
            unsafe_allow_html=True,
        )
        ov1, ov2 = st.columns([1, 3])
        cp_pct_override = ov1.number_input(
            "% Complete Override",
            min_value=0, max_value=100, step=1,
            value=int(float(edit_proj.get("pct_override", 0) or 0)),
            key=f"cp_pct_override_{eid}",
            label_visibility="collapsed",
        )
        if cp_pct_override > 0:
            ov2.info(
                f"Cards will show **{cp_pct_override}%** instead of the "
                f"auto-calculated value."
            )
        else:
            ov2.caption("Currently auto-calculating from milestone dates.")

        def _fmt_date(d):
            return d.isoformat() if d else ""

        # ── Save / Delete buttons ─────────────────────────────────────────────
        save_label = "Update Project" if edit_id else "Add Project"
        if edit_id:
            btn_save_col, btn_del_col = st.columns([3, 1])
            do_save   = btn_save_col.button(save_label, type="primary", use_container_width=True)
            do_delete = btn_del_col.button("Delete", type="secondary", use_container_width=True)
        else:
            do_save   = st.button(save_label, type="primary", use_container_width=True)
            do_delete = False

        if do_delete and edit_id:
            updated = projects[projects["id"] != edit_id]
            custom_pillars = [p for p in all_pillars if p not in DEFAULT_PILLARS]
            extra_codes    = {k: v for k, v in all_codes.items()
                              if k not in CONTENT_ACCT_CODES}
            ok, msg = save_content_projects(
                updated,
                {"rtl_hourly_rate": rate},
                custom_pillars,
                extra_codes,
            )
            if ok:
                st.session_state.pop("cp_edit_sel", None)
                st.toast(f"Deleted '{edit_proj.get('title', '')}'.")
                st.rerun()
            else:
                st.error(msg)

        if do_save:
            if not cp_title.strip():
                st.error("Project title is required.")
            else:
                cp_code = cp_code_raw.split(" — ")[0].strip() if cp_code_raw else ""
                new_row = {
                    "id":                   edit_id if edit_id else (
                        int(projects["id"].max() + 1)
                        if not projects.empty and projects["id"].notna().any() else 1
                    ),
                    "title":                cp_title.strip(),
                    "type":                 cp_type,
                    "pillar":               cp_pillar,
                    "acct_code":            cp_code,
                    "owner":                cp_owner,
                    "sponsored":            cp_sponsored,
                    "sponsorship_type":     cp_sponsorship_type,
                    "sponsorship_other":    cp_sponsorship_other,
                    "status":               cp_status,
                    "format":               cp_format,
                    "content_generator":    cp_generator,
                    "committee_name":       cp_committee,
                    "funding_source":       cp_funding,
                    "program_partner_name": cp_partner_name,
                    "budget":               cp_budget,
                    "est_hours":            cp_hours,
                    "draft_delivered":      _fmt_date(cp_draft_del),
                    "draft_commented":      _fmt_date(cp_draft_com),
                    "draft_completed":      _fmt_date(cp_draft_cpd),
                    "layout1_delivered":    _fmt_date(cp_lay1_del),
                    "layout1_commented":    _fmt_date(cp_lay1_com),
                    "layout2_delivered":    _fmt_date(cp_lay2_del),
                    "layout2_approved":     _fmt_date(cp_lay2_apr),
                    "print_date":           _fmt_date(cp_print_date) if cp_print_date else "",
                    "go_live_date":         _fmt_date(cp_go_live)    if cp_go_live    else "",
                    "notes":                cp_notes,
                    "pct_override":         cp_pct_override,
                    "confirmed_pending":    cp_conf_pend,
                }
                new_df = pd.DataFrame([new_row])
                if edit_id:
                    updated = projects[projects["id"] != edit_id]
                    updated = pd.concat([updated, new_df], ignore_index=True)
                else:
                    updated = pd.concat([projects, new_df], ignore_index=True)

                custom_pillars = [p for p in all_pillars if p not in DEFAULT_PILLARS]
                extra_codes    = {k: v for k, v in all_codes.items()
                                  if k not in CONTENT_ACCT_CODES}
                ok, msg = save_content_projects(
                    updated,
                    {"rtl_hourly_rate": rate},
                    custom_pillars,
                    extra_codes,
                )
                if ok:
                    # Reset selector to "Add new project" after a successful save
                    st.session_state.pop("cp_edit_sel", None)
                    st.toast("Project saved.")
                    st.rerun()
                else:
                    st.error(msg)

    # ── Dashboard ─────────────────────────────────────────────────────────────
    if projects.empty:
        st.info("No projects yet — use the form above to add your first project.")
        return

    # ── Key Metrics ───────────────────────────────────────────────────────────
    st.markdown("##### Key Metrics")

    # Confirmed-only subset (handle projects saved before confirmed_pending existed)
    def _is_confirmed(row):
        return str(row.get("confirmed_pending", "Confirmed") or "Confirmed") == "Confirmed"

    confirmed = projects[projects.apply(_is_confirmed, axis=1)]
    conf_pcts  = (confirmed.apply(_compute_project_pct, axis=1)
                  if not confirmed.empty else pd.Series(dtype=float))
    avg_pct    = round(conf_pcts.mean()) if not conf_pcts.empty else 0

    pillar_counts = projects["pillar"].value_counts()
    active_pillars = [p for p in all_pillars if pillar_counts.get(p, 0) > 0]

    km1, km2, km3 = st.columns(3)

    with km1:
        st.metric("Pieces of Content", len(projects))

    with km2:
        pillar_lines = "".join(
            f"<div style='margin:3px 0;font-size:0.8rem'>"
            f"<span style='color:{_pillar_color(p)}'>\u25cf</span>"
            f" <span style='color:#FCFCFC'>{p}</span>"
            f"<span style='color:#9E9E9E'>: {int(pillar_counts.get(p, 0))}</span>"
            f"</div>"
            for p in active_pillars
        )
        st.markdown(
            "<div style='background:#282828;border-radius:6px;padding:12px 16px;"
            "border-left:3px solid #B4E817'>"
            "<div style='font-size:0.85rem;color:#FCFCFC;margin-bottom:6px'>"
            "Content by Pillar</div>"
            f"{pillar_lines}"
            "</div>",
            unsafe_allow_html=True,
        )

    with km3:
        st.metric("Avg % Complete", f"{avg_pct}%",
                  help="Confirmed projects only")

    st.divider()

    # ── Go-Live Timeline ──────────────────────────────────────────────────────
    st.markdown("##### Go-Live Timeline")
    tl_key_html = "<div style='display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px'>"
    for pillar in all_pillars:
        col = _pillar_color(pillar)
        tl_key_html += (
            f"<span style='background:{col}22;color:{col};"
            f"border:1px solid {col}55;border-radius:10px;"
            f"font-size:0.75rem;font-weight:600;padding:3px 10px'>{pillar}</span>"
        )
    tl_key_html += "</div>"
    st.markdown(tl_key_html, unsafe_allow_html=True)

    year = date.today().year

    def _parse_gl(v):
        if not v or str(v) in ("", "None", "NaT", "nan"):
            return None
        try:
            return date.fromisoformat(str(v)[:10])
        except (ValueError, TypeError):
            return None

    tl = projects.copy()
    tl["_gl"] = tl["go_live_date"].apply(_parse_gl)
    tl = tl[tl["_gl"].apply(lambda d: d is not None and d.year == year)]

    if tl.empty:
        st.caption("No go-live dates in the current year to display.")
    else:
        year_start  = date(year, 1, 1)
        tl["_week"] = tl["_gl"].apply(lambda d: (d - year_start).days // 7)
        tl_pillars  = [p for p in all_pillars if p in tl["pillar"].values]

        # Max items stacked per (pillar, week) — used for dynamic row height
        stacks_df  = (tl.groupby(["pillar", "_week"]).size()
                        .reset_index(name="cnt"))
        max_per_p  = stacks_df.groupby("pillar")["cnt"].max()

        # Assign base Y for each pillar, expanding row height for stacks
        y_bases = {}
        cur_y   = 0.0
        for p in tl_pillars:
            y_bases[p] = cur_y
            ms   = int(max_per_p.get(p, 1))
            cur_y += max(1.0, ms * 0.45 + 0.3)

        total_y = cur_y
        chart_h = max(220, int(total_y * 65) + 80)

        tl = tl.sort_values("_gl")
        tl["_stack"] = tl.groupby(["pillar", "_week"]).cumcount()
        tl["_y"]     = tl.apply(
            lambda r: y_bases.get(str(r["pillar"]), 0) + r["_stack"] * 0.45,
            axis=1,
        )

        fig_tl = go.Figure()
        for pillar in tl_pillars:
            pdf      = tl[tl["pillar"] == pillar]
            pc       = _pillar_color(pillar)
            pct_lbl  = pdf.apply(_compute_project_pct, axis=1).astype(str) + "% complete"
            conf_lbl = pdf.apply(
                lambda r: str(r.get("confirmed_pending", "Confirmed") or "Confirmed"),
                axis=1,
            ).tolist()
            opacities = [0.5 if c == "Pending" else 1.0 for c in conf_lbl]
            fig_tl.add_trace(go.Scatter(
                x=pdf["_gl"].apply(lambda d: d.isoformat()),
                y=pdf["_y"],
                mode="markers",
                name=pillar,
                marker=dict(symbol="square", size=18, color=pc, opacity=opacities),
                customdata=list(zip(pdf["title"].tolist(), pct_lbl.tolist(), conf_lbl)),
                hovertemplate=(
                    "<b>%{customdata[0]}</b><br>"
                    "Go-live: %{x|%Y-%m-%d}<br>"
                    "%{customdata[1]} · %{customdata[2]}<extra></extra>"
                ),
                showlegend=False,
            ))

        tick_vals = [y_bases[p] for p in tl_pillars]
        tick_text = [
            p.replace("Advertising and Sales-Driven", "Advert. & Sales-Driven")
            for p in tl_pillars
        ]

        fig_tl.update_layout(**_chart_base(
            height=chart_h,
            xaxis=dict(
                range=[f"{year}-01-01", f"{year}-12-31"],
                tickformat="%b",
                dtick="M1",
                tickfont=dict(color=c_colors["subtext"], size=11,
                              family="Inter, Arial, sans-serif"),
                gridcolor=c_colors["grid"],
                zerolinecolor=c_colors["grid"],
                showgrid=True,
            ),
            yaxis=dict(
                tickvals=tick_vals,
                ticktext=tick_text,
                tickfont=dict(color=c_colors["text"], size=11,
                              family="Inter, Arial, sans-serif"),
                showgrid=False,
                range=[-0.4, total_y],
                automargin=True,
            ),
            showlegend=False,
            margin=dict(l=10, r=20, t=20, b=40),
        ))

        st.plotly_chart(fig_tl, use_container_width=True)

    st.divider()

    # ── % Complete by Pillar (Confirmed) ─────────────────────────────────────
    st.markdown("##### % Complete by Pillar — Confirmed Projects")

    pillar_avgs = {
        p: round(confirmed[confirmed["pillar"] == p]
                 .apply(_compute_project_pct, axis=1).mean())
        for p in all_pillars
        if not confirmed.empty and not confirmed[confirmed["pillar"] == p].empty
    }

    if not pillar_avgs:
        st.caption("No confirmed projects yet.")
    else:
        donut_items = list(pillar_avgs.items())
        for row_start in range(0, len(donut_items), 3):
            d_cols = st.columns(3)
            for ci, (pillar, avg) in enumerate(donut_items[row_start:row_start + 3]):
                pc = _pillar_color(pillar)
                with d_cols[ci]:
                    fig_d = go.Figure(go.Pie(
                        values=[avg, max(0, 100 - avg)],
                        labels=["Complete", "Remaining"],
                        hole=0.60,
                        marker_colors=[pc, "#3A3A3A"],
                        hoverinfo="skip",
                        textinfo="none",
                        sort=False,
                    ))
                    fig_d.add_annotation(
                        text=f"<b>{avg}%</b>",
                        x=0.5, y=0.5,
                        font=dict(size=22, color=c_colors["text"],
                                  family="Inter, Arial, sans-serif"),
                        showarrow=False,
                    )
                    short = pillar if len(pillar) <= 24 else pillar[:22] + "\u2026"
                    fig_d.update_layout(**_chart_base(
                        height=200,
                        title=dict(
                            text=short,
                            font=dict(color=pc, size=11,
                                      family="Inter, Arial, sans-serif"),
                            x=0.5, xanchor="center",
                            y=0.97, yanchor="top",
                        ),
                        showlegend=False,
                        margin=dict(t=32, b=10, l=10, r=10),
                    ))
                    st.plotly_chart(fig_d, use_container_width=True)

    st.divider()

    # Pillar color key
    st.markdown("##### Pillar Key")
    key_html = "<div style='display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px'>"
    for pillar in all_pillars:
        col = _pillar_color(pillar)
        key_html += (
            f"<span style='background:{col}22;color:{col};"
            f"border:1px solid {col}55;border-radius:10px;"
            f"font-size:0.75rem;font-weight:600;padding:3px 10px'>{pillar}</span>"
        )
    key_html += "</div>"
    st.markdown(key_html, unsafe_allow_html=True)

    # Filter + sort bar
    st.markdown("##### Projects")
    fa, fb, fc, fd = st.columns(4)
    f_pillar = fa.selectbox("Filter: Pillar",  ["All"] + all_pillars,      key="cp_f_pillar")
    f_status = fb.selectbox("Filter: Priority",  ["All"] + PROJECT_STATUSES, key="cp_f_status")
    f_owner  = fc.selectbox("Filter: Owner",   ["All"] + RTL_OWNERS,       key="cp_f_owner")
    sort_by  = fd.selectbox(
        "Sort by",
        ["Date added", "% Complete (high–low)", "% Complete (low–high)",
         "Go-live (soonest)", "Go-live (latest)"],
        key="cp_sort_by",
    )

    mask = pd.Series(True, index=projects.index)
    if f_pillar != "All": mask &= projects["pillar"] == f_pillar
    if f_status != "All": mask &= projects["status"] == f_status
    if f_owner  != "All": mask &= projects["owner"]  == f_owner

    visible = projects[mask].copy()

    if sort_by == "% Complete (high–low)":
        visible["_pct"] = visible.apply(_compute_project_pct, axis=1)
        visible = visible.sort_values("_pct", ascending=False)
    elif sort_by == "% Complete (low–high)":
        visible["_pct"] = visible.apply(_compute_project_pct, axis=1)
        visible = visible.sort_values("_pct", ascending=True)
    elif sort_by in ("Go-live (soonest)", "Go-live (latest)"):
        visible["_go_live_sort"] = pd.to_datetime(
            visible["go_live_date"], errors="coerce"
        )
        visible = visible.sort_values(
            "_go_live_sort",
            ascending=(sort_by == "Go-live (soonest)"),
            na_position="last",
        )

    visible = visible.reset_index(drop=True)
    if visible.empty:
        st.info("No projects match the current filters.")
        return

    # 2-column card grid
    for row_start in range(0, len(visible), 2):
        cols = st.columns(2)
        for col_idx in range(2):
            proj_idx = row_start + col_idx
            if proj_idx >= len(visible):
                break
            proj = visible.iloc[proj_idx].to_dict()
            pc   = _pillar_color(proj.get("pillar", ""))
            with cols[col_idx]:
                st.markdown(_project_card_html(proj, pc), unsafe_allow_html=True)

    # ── Global Settings (bottom of page) ─────────────────────────────────────
    st.divider()
    with st.expander("Global Settings", expanded=False):
        gs1, gs2 = st.columns(2)
        hourly_rate = gs1.number_input(
            "RTL Staff Cost ($/hr)",
            min_value=0.0, step=0.01,
            value=default_rate,
            format="%.2f",
            key="cp_hourly_rate",
            help="Used to calculate estimated RTL labor cost per project. "
                 "Auto-filled from Financial KPIs if set.",
        )
        if auto_rate > 0:
            gs1.caption(f"Derived from Financial KPIs annual cost: ${auto_rate:.2f}/hr")

        # ── Add custom pillar ─────────────────────────────────────────────────
        st.markdown(
            "<p style='color:#9E9E9E;font-size:0.85rem;margin-top:10px'>"
            "Add custom pillar (leave blank to skip)</p>",
            unsafe_allow_html=True,
        )
        pc1, _ = st.columns([3, 1])
        new_pillar = pc1.text_input("New pillar name", key="cp_new_pillar",
                                    label_visibility="collapsed",
                                    placeholder="e.g. Joint Venture")

        # ── Add custom accounting code ────────────────────────────────────────
        st.markdown(
            "<p style='color:#9E9E9E;font-size:0.85rem;margin-top:10px'>"
            "Add custom accounting code (leave blank to skip)</p>",
            unsafe_allow_html=True,
        )
        cc1, cc2, _ = st.columns([1, 3, 1])
        new_code_id   = cc1.text_input("Code", key="cp_new_code_id",
                                       label_visibility="collapsed",
                                       placeholder="e.g. 711")
        new_code_name = cc2.text_input("Code name", key="cp_new_code_name",
                                       label_visibility="collapsed",
                                       placeholder="e.g. Special Projects")

        if st.button("Save Global Settings", type="primary"):
            new_settings   = {"rtl_hourly_rate": hourly_rate}
            custom_pillars = [p for p in all_pillars if p not in DEFAULT_PILLARS]
            extra_codes    = {k: v for k, v in all_codes.items() if k not in CONTENT_ACCT_CODES}

            if new_pillar.strip() and new_pillar.strip() not in all_pillars:
                custom_pillars.append(new_pillar.strip())
            if new_code_id.strip() and new_code_name.strip():
                extra_codes[new_code_id.strip()] = new_code_name.strip()

            # Clear cached auto-rate so it re-derives on next load
            st.session_state.pop("cp_auto_rate", None)
            ok, msg = save_content_projects(projects, new_settings, custom_pillars, extra_codes)
            if ok:
                st.toast("Settings saved.")
                st.rerun()
            else:
                st.error(msg)


# ══════════════════════════════════════════════════════════════════════════════
# AUTH — simple password gate
# ══════════════════════════════════════════════════════════════════════════════

def _hide_sidebar():
    st.markdown(
        "<style>section[data-testid='stSidebar']{display:none}</style>",
        unsafe_allow_html=True,
    )


def _check_auth() -> bool:
    """
    Show a password screen if not yet authenticated.

    Two passwords are supported:
      APP_PASSWORD   → staff role  (landing page → time entry)
      ADMIN_PASSWORD → admin role  (straight to dashboard, read-only)

    Both fall back to local defaults if not set in st.secrets.
    Returns True once authenticated.
    """
    if st.session_state.get("authenticated"):
        return True

    _hide_sidebar()

    _, col, _ = st.columns([1, 1.6, 1])
    with col:
        st.markdown("<div style='padding-top:80px'></div>", unsafe_allow_html=True)
        st.markdown(
            "<h2 style='color:#FCFCFC;font-family:Inter,Arial,sans-serif;"
            "font-weight:600;margin-bottom:4px'>RTL KPI Tracker</h2>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "<p style='color:#9E9E9E;font-family:Inter,Arial,sans-serif;"
            "margin-bottom:28px'>Council on Vertical Urbanism</p>",
            unsafe_allow_html=True,
        )
        pwd = st.text_input("Password", type="password", placeholder="Enter password",
                            label_visibility="collapsed")
        if st.button("Sign In", type="primary", use_container_width=True):
            try:
                staff_pwd = st.secrets["APP_PASSWORD"]
            except Exception:
                staff_pwd = "rtl2026"
            try:
                admin_pwd = st.secrets["ADMIN_PASSWORD"]
            except Exception:
                admin_pwd = "rtladmin"

            if pwd == staff_pwd:
                st.session_state["authenticated"] = True
                st.session_state["role"] = "staff"
                st.rerun()
            elif pwd == admin_pwd:
                st.session_state["authenticated"] = True
                st.session_state["role"] = "admin"
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# LANDING PAGE — staff member selection
# ══════════════════════════════════════════════════════════════════════════════

def view_landing():
    """Welcome screen — staff pick their name before entering time."""
    _hide_sidebar()

    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown("<div style='padding-top:60px'></div>", unsafe_allow_html=True)
        st.markdown(
            "<h2 style='color:#FCFCFC;font-family:Inter,Arial,sans-serif;"
            "font-weight:600;margin-bottom:6px'>Who are you?</h2>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "<p style='color:#9E9E9E;font-family:Inter,Arial,sans-serif;"
            "margin-bottom:32px'>Select your name to start entering your time.</p>",
            unsafe_allow_html=True,
        )

        names = list(STAFF.keys())
        row1, row2 = st.columns(2), st.columns(2)
        for i, name in enumerate(names):
            with (row1 if i < 2 else row2)[i % 2]:
                st.markdown(
                    f"<div style='background:#282828;border:1px solid #4F4F4F;"
                    f"border-left:4px solid #B4E817;border-radius:6px;"
                    f"padding:18px 16px;margin-bottom:4px;'>"
                    f"<span style='color:#FCFCFC;font-family:Inter,Arial,sans-serif;"
                    f"font-weight:600;font-size:1rem'>{name}</span></div>",
                    unsafe_allow_html=True,
                )
                if st.button("Select", key=f"land_{name}", use_container_width=True):
                    st.session_state["person"] = name
                    st.rerun()

        st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)
        st.markdown(
            "<p style='color:#4F4F4F;font-size:0.75rem;text-align:center;"
            "font-family:Inter,Arial,sans-serif'>Manager? Use the Team Overview "
            "in the sidebar after selecting any name.</p>",
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# VIEW: DATABASE KPIs
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner=False)
def _get_mysql_conn():
    """Return a live mysql-connector connection using secrets, or None on error."""
    try:
        import mysql.connector
        try:
            cfg = st.secrets.get("mysql", {})
        except Exception:
            cfg = {}
        return mysql.connector.connect(
            host       = cfg.get("host", "localhost"),
            port       = int(cfg.get("port", 3306)),
            database   = cfg.get("database", ""),
            user       = cfg.get("user", ""),
            password   = cfg.get("password", ""),
            connection_timeout = 10,
            use_pure   = True,   # avoids C-extension malloc crash on Apple Silicon / Py 3.13
        )
    except Exception:
        return None


@st.cache_data(ttl=3600, show_spinner=False)
def _bldg_query(sql: str) -> pd.DataFrame:
    """Run a read-only MySQL query and return a DataFrame. Cached 1 hr."""
    conn = _get_mysql_conn()
    if conn is None:
        return pd.DataFrame()
    try:
        if not conn.is_connected():
            conn.reconnect(attempts=2, delay=1)
        cursor = conn.cursor(dictionary=True)
        cursor.execute(sql)
        rows = cursor.fetchall()
        cursor.close()
        return pd.DataFrame(rows) if rows else pd.DataFrame()
    except Exception as e:
        st.error(f"Database query failed: {e}")
        return pd.DataFrame()


def view_building_kpis():
    c = _chart_colors()

    # ── Credential check ──────────────────────────────────────────────────────
    try:
        _has_mysql = "mysql" in st.secrets
    except Exception:
        _has_mysql = False

    if not _has_mysql:
        st.warning(
            "MySQL credentials not configured. Add a `[mysql]` block to "
            "`.streamlit/secrets.toml` (local) or to Streamlit Cloud Secrets.",
        )
        st.code(
            "[mysql]\nhost     = \"your-host\"\nport     = 3306\n"
            "database = \"your-database\"\nuser     = \"your-user\"\n"
            "password = \"your-password\"",
            language="toml",
        )
        return

    START = "2018-01-01"
    FIRST_YEAR = 2018
    current_year = date.today().year
    jan1_cy = f"{current_year}-01-01"

    # ══════════════════════════════════════════════════════════════════════════
    # KEY METRICS
    # ══════════════════════════════════════════════════════════════════════════
    df_totals = _bldg_query(f"""
        SELECT
          COUNT(*) AS total,
          SUM(CASE WHEN b.date_create >= '{jan1_cy}' THEN 1 ELSE 0 END) AS added_cy,
          SUM(CASE WHEN b.date_update >= '{jan1_cy}'
                    AND (b.date_create IS NULL OR b.date_create < '{jan1_cy}')
               THEN 1 ELSE 0 END) AS updated_cy
        FROM ctbuh_building b
        WHERE b.deleted_at IS NULL
    """)
    df_ctry_cnt = _bldg_query("""
        SELECT COUNT(DISTINCT b.country_id) AS n
        FROM   ctbuh_building b
        WHERE  b.deleted_at IS NULL
          AND  b.country_id IS NOT NULL AND b.country_id != ''
    """)
    if not df_totals.empty:
        total_bldgs = int(df_totals["total"].iloc[0])
        added_cy    = int(df_totals["added_cy"].iloc[0])
        updated_cy  = int(df_totals["updated_cy"].iloc[0])
        ctry_cnt    = int(df_ctry_cnt["n"].iloc[0]) if not df_ctry_cnt.empty else 0
        km1, km2, km3, km4 = st.columns(4)
        km1.metric("Total Buildings",          f"{total_bldgs:,}")
        km2.metric(f"Added in {current_year}", f"{added_cy:,}")
        km3.metric(f"Updated in {current_year} (existing)", f"{updated_cy:,}")
        km4.metric("Countries Represented",    f"{ctry_cnt:,}")

    # ══════════════════════════════════════════════════════════════════════════
    # PROGRESS TO GOAL
    # ══════════════════════════════════════════════════════════════════════════
    ANNUAL_GOALS = {
        2026: 47_250,
        2027: 49_000,
        2028: 51_000,
        2029: 53_250,
        2030: 55_750,
    }
    if current_year in ANNUAL_GOALS:
        goal_target = ANNUAL_GOALS[current_year]

        df_jan1 = _bldg_query(f"""
            SELECT COUNT(*) AS total
            FROM   ctbuh_building b
            WHERE  b.deleted_at IS NULL
              AND  b.date_create <= '{jan1_cy}'
        """)
        df_goal = _bldg_query("""
            SELECT COUNT(*) AS total
            FROM   ctbuh_building b
            WHERE  b.deleted_at IS NULL
        """)

        if not df_goal.empty and not df_jan1.empty:
            jan1_count    = int(df_jan1["total"].iloc[0])
            current_total = int(df_goal["total"].iloc[0])
            remaining     = max(goal_target - current_total, 0)
            year_gap      = goal_target - jan1_count
            gained        = current_total - jan1_count
            pct           = round(gained / year_gap * 100, 1) if year_gap > 0 else 100.0

            goal_col1, goal_col2 = st.columns([1, 2])
            with goal_col1:
                fig_goal = go.Figure(go.Pie(
                    labels=["Progress this year", "Remaining to goal"],
                    values=[gained, max(year_gap - gained, 0)],
                    marker=dict(colors=[CVU_GREEN, CVU_SURFACE]),
                    textinfo="none",
                    hovertemplate="<b>%{label}</b><br>%{value:,}<extra></extra>",
                    hole=0.5,
                ))
                fig_goal.add_annotation(
                    text=f"<b>{pct}%</b>",
                    x=0.5, y=0.5,
                    font=dict(size=22, color=c["text"]),
                    showarrow=False,
                )
                fig_goal.update_layout(**_chart_base(
                    title=dict(
                        text=f"{current_year} Progress to Goal ({goal_target:,} buildings)",
                        font=dict(size=14, color=c["text"]),
                    ),
                    height=340,
                    legend=dict(bgcolor="rgba(0,0,0,0)",
                                font=dict(color=c["text"], size=11)),
                    showlegend=True,
                ))
                st.plotly_chart(fig_goal, use_container_width=True,
                                config={"displayModeBar": False})
            with goal_col2:
                st.markdown(
                    f"""
                    <div style='padding:24px 0 0 16px; color:{c["text"]}'>
                        <p style='font-size:15px; margin-bottom:6px'>
                            <b>Buildings as of Jan 1, {current_year}:</b> {jan1_count:,}
                        </p>
                        <p style='font-size:15px; margin-bottom:6px'>
                            <b>Current buildings:</b> {current_total:,}
                        </p>
                        <p style='font-size:15px; margin-bottom:6px'>
                            <b>{current_year} target:</b> {goal_target:,}
                        </p>
                        <p style='font-size:15px; margin-bottom:6px'>
                            <b>Remaining:</b> {remaining:,}
                        </p>
                        <p style='font-size:18px; margin-top:12px'>
                            <b>{pct}% to goal</b>
                        </p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
    else:
        st.info(f"No goal defined for {current_year}.")

    st.divider()

    # ── Controls: filter mode + refresh ──────────────────────────────────────
    ctrl_col, _ = st.columns([3, 4])

    filter_mode = ctrl_col.radio(
        "Filter by:",
        ["All years", "Year", "Custom date range"],
        horizontal=True,
        key="bldg_filter_mode",
    )

    selected_year = None
    custom_start = None
    custom_end   = None

    if filter_mode == "Year":
        year_options = [str(y) for y in range(FIRST_YEAR, current_year + 1)]
        selected_year = ctrl_col.selectbox(
            "Select year:", year_options,
            index=len(year_options) - 1,
            key="bldg_year_filter",
        )

    elif filter_mode == "Custom date range":
        date_col1, date_col2 = ctrl_col.columns(2)
        custom_start = date_col1.date_input(
            "From:", value=date(FIRST_YEAR, 1, 1),
            min_value=date(FIRST_YEAR, 1, 1), max_value=date.today(),
            key="bldg_custom_start",
        )
        custom_end = date_col2.date_input(
            "To:", value=date.today(),
            min_value=date(FIRST_YEAR, 1, 1), max_value=date.today(),
            key="bldg_custom_end",
        )
        if custom_start > custom_end:
            ctrl_col.warning("'From' date must be before 'To' date.")
            custom_start, custom_end = custom_end, custom_start

    if ctrl_col.button("Refresh Database Data"):
        _bldg_query.clear()
        _get_mysql_conn.clear()
        st.rerun()

    # Build SQL date range clauses based on selection
    if filter_mode == "All years":
        filt_b        = f"b.date_create >= '{START}'"
        filt_b_update = "1=1"
        filt_h        = f"h.created_at  >= '{START}'"
        filt_b_pie    = "1=1"
        filt_reno_date = (
            f"GREATEST(COALESCE(b.date_create,'1900-01-01'),"
            f"COALESCE(b.date_update,'1900-01-01')) >= '{START}'"
        )
    elif filter_mode == "Year":
        yr = selected_year
        filt_b        = f"b.date_create BETWEEN '{yr}-01-01' AND '{yr}-12-31'"
        filt_b_update = f"b.date_update BETWEEN '{yr}-01-01' AND '{yr}-12-31'"
        filt_h        = f"h.created_at  BETWEEN '{yr}-01-01' AND '{yr}-12-31'"
        filt_b_pie    = f"b.date_create BETWEEN '{yr}-01-01' AND '{yr}-12-31'"
        filt_reno_date = (
            f"GREATEST(COALESCE(b.date_create,'1900-01-01'),"
            f"COALESCE(b.date_update,'1900-01-01'))"
            f" BETWEEN '{yr}-01-01' AND '{yr}-12-31'"
        )
    else:  # Custom date range
        s = custom_start.strftime("%Y-%m-%d")
        e = custom_end.strftime("%Y-%m-%d")
        filt_b        = f"b.date_create BETWEEN '{s}' AND '{e}'"
        filt_b_update = f"b.date_update BETWEEN '{s}' AND '{e}'"
        filt_h        = f"h.created_at  BETWEEN '{s}' AND '{e}'"
        filt_b_pie    = f"b.date_create BETWEEN '{s}' AND '{e}'"
        filt_reno_date = (
            f"GREATEST(COALESCE(b.date_create,'1900-01-01'),"
            f"COALESCE(b.date_update,'1900-01-01'))"
            f" BETWEEN '{s}' AND '{e}'"
        )

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — OVERALL DATABASE
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### Overall Database")

    # ── Helper: build a standard monthly line chart ───────────────────────────
    def _monthly_line(df, y_col, title, color=CVU_GREEN):
        if df.empty:
            st.info(f"No data returned for: {title}")
            return
        df = df.copy()
        df["month"] = pd.to_datetime(df["month"] + "-01")
        df = df.sort_values("month")
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df["month"], y=df[y_col],
            mode="lines+markers",
            line=dict(color=color, width=2),
            marker=dict(size=4, color=color),
            hovertemplate="%{x|%b %Y}: <b>%{y:,}</b><extra></extra>",
        ))
        fig.update_layout(**_chart_base(
            title=dict(text=title, font=dict(size=14, color=c["text"])),
            height=280,
            xaxis=dict(
                gridcolor=c["grid"], zerolinecolor=c["grid"],
                tickfont=dict(color=c["subtext"], size=11),
                tickformat="%b %Y",
            ),
            yaxis=dict(
                gridcolor=c["grid"], zerolinecolor=c["grid"],
                tickfont=dict(color=c["subtext"], size=11),
                tickformat=",d",
            ),
        ))
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    # ── 1. New buildings added ────────────────────────────────────────────────
    # NOTE: ctbuh_building uses date_create / date_update (not created_at / updated_at)
    df_new_bldg = _bldg_query(f"""
        SELECT DATE_FORMAT(b.date_create, '%Y-%m') AS month,
               COUNT(*) AS count
        FROM   ctbuh_building b
        WHERE  b.deleted_at IS NULL
          AND  {filt_b}
        GROUP  BY month
        ORDER  BY month
    """)

    # ── 2. Building updates ───────────────────────────────────────────────────
    df_updates = _bldg_query(f"""
        SELECT DATE_FORMAT(h.created_at, '%Y-%m') AS month,
               COUNT(*) AS count
        FROM   history h
        WHERE  h.model_type = 'Building'
          AND  h.type = 'Update'
          AND  {filt_h}
          AND  DATE_FORMAT(h.created_at, '%Y-%m') NOT IN ('2021-06', '2022-10')
        GROUP  BY month
        ORDER  BY month
    """)

    # ── 3. New images ─────────────────────────────────────────────────────────
    df_images = _bldg_query(f"""
        SELECT DATE_FORMAT(h.created_at, '%Y-%m') AS month,
               COUNT(*) AS count
        FROM   history h
        WHERE  h.model_type = 'Image'
          AND  h.type = 'Create'
          AND  {filt_h}
        GROUP  BY month
        ORDER  BY month
    """)

    # ── 4. New company connections ────────────────────────────────────────────
    df_companies = _bldg_query(f"""
        SELECT DATE_FORMAT(h.created_at, '%Y-%m') AS month,
               SUM(CASE WHEN h.type = 'attach_company' THEN 1 ELSE -1 END) AS count
        FROM   history h
        WHERE  h.model_type = 'Building'
          AND  h.type IN ('attach_company', 'detach_company')
          AND  {filt_h}
        GROUP  BY month
        ORDER  BY month
    """)

    col1, col2 = st.columns(2)
    with col1:
        _monthly_line(df_new_bldg,  "count", "New Buildings Added")
        _monthly_line(df_images,    "count", "New Images Added")
    with col2:
        _monthly_line(df_updates,   "count", "Building Updates")
        _monthly_line(df_companies, "count", "New Company Connections")

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2 — BUILDINGS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### Buildings")

    # ── 5. Buildings by function (pie) ────────────────────────────────────────
    # main_use_01/02 are enums — use exact value matching
    df_func = _bldg_query(f"""
        SELECT
          CASE
            WHEN b.main_use_02 != ''
                 THEN 'Mixed-Use'
            WHEN b.main_use_01 = 'office'
                 THEN 'All-Office'
            WHEN b.main_use_01 = 'residential'
                 THEN 'All-Residential'
            WHEN b.main_use_01 = 'hotel'
                 THEN 'All-Hotel'
            ELSE 'Other'
          END AS function_group,
          COUNT(*) AS count
        FROM   ctbuh_building b
        WHERE  b.deleted_at IS NULL
          AND  {filt_b_pie}
        GROUP  BY function_group
        ORDER  BY count DESC
    """)

    # ── 6. Buildings by structural material (pie) ─────────────────────────────
    # structural_material is an enum — use exact value matching
    df_mat = _bldg_query(f"""
        SELECT
          CASE b.structural_material
            WHEN 'steel'                     THEN 'All-Steel'
            WHEN 'concrete'                  THEN 'All-Concrete'
            WHEN 'composite'                 THEN 'Composite'
            WHEN 'timber'                    THEN 'Timber'
            WHEN 'timber/concrete'           THEN 'Timber'
            WHEN 'timber/composite'          THEN 'Timber'
            WHEN 'timber composite/concrete' THEN 'Timber'
            WHEN 'concrete/steel'            THEN 'Mixed'
            WHEN 'steel/concrete'            THEN 'Mixed'
            ELSE 'Other/Unknown'
          END AS material_group,
          COUNT(*) AS count
        FROM   ctbuh_building b
        WHERE  b.deleted_at IS NULL
          AND  {filt_b_pie}
        GROUP  BY material_group
        ORDER  BY count DESC
    """)

    # ── 7. Renovations & Retrofits over time (line) ───────────────────────────
    df_reno = _bldg_query(f"""
        SELECT DATE_FORMAT(
                 GREATEST(
                   COALESCE(b.date_create, '1900-01-01'),
                   COALESCE(b.date_update, '1900-01-01')
                 ), '%Y-%m'
               ) AS month,
               COUNT(DISTINCT b.id) AS count
        FROM   ctbuh_building b
        LEFT JOIN ctbuh_building lb
               ON lb.id = b.linked_building_id
              AND lb.deleted_at IS NULL
        WHERE  b.deleted_at IS NULL
          AND  (
                 (b.retrofit_start IS NOT NULL AND b.retrofit_start != 0)
              OR (b.retrofit_end   IS NOT NULL AND b.retrofit_end   != 0)
              OR (b.recladding_year IS NOT NULL AND b.recladding_year != 0)
              OR UPPER(TRIM(b.linked_building_status)) = 'RENOVATED'
              OR UPPER(TRIM(b.status)) = 'UREN'
              OR UPPER(TRIM(lb.status)) = 'UREN'
          )
          AND  {filt_reno_date}
        GROUP  BY month
        ORDER  BY month
    """)

    # ── Pie chart helper ──────────────────────────────────────────────────────
    def _pie(df, label_col, value_col, title):
        if df.empty:
            st.info(f"No data for: {title}")
            return
        PALETTE = [CVU_GREEN] + CVU_PALETTE + [CVU_GRAY]
        fig = go.Figure(go.Pie(
            labels=df[label_col],
            values=df[value_col],
            marker=dict(colors=PALETTE[:len(df)]),
            textfont=dict(color="#171717", size=12),
            hovertemplate="<b>%{label}</b><br>%{value:,} (%{percent})<extra></extra>",
            hole=0.35,
        ))
        fig.update_layout(**_chart_base(
            title=dict(text=title, font=dict(size=14, color=c["text"])),
            height=340,
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=c["text"], size=11)),
            showlegend=True,
        ))
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    pie_col1, pie_col2 = st.columns(2)
    with pie_col1:
        _pie(df_func, "function_group", "count", "Buildings by Function")
    with pie_col2:
        _pie(df_mat,  "material_group", "count", "Buildings by Structural Material")

    _monthly_line(df_reno, "count", "Renovations & Retrofits Added / Updated Over Time",
                  color=CVU_PALETTE[2])

    # ── 8. Buildings by Status (bar) ─────────────────────────────────────────
    df_status = _bldg_query(f"""
        SELECT
          CASE b.status
            WHEN 'COM'  THEN 'Complete'
            WHEN 'UCT'  THEN 'Under Construction'
            WHEN 'STO'  THEN 'Under Construction'
            WHEN 'UC'   THEN 'Under Construction'
            WHEN 'PRO'  THEN 'Proposed'
            WHEN 'DEM'  THEN 'Demolished'
            WHEN 'UDEM' THEN 'Demolished'
            WHEN 'REN'  THEN 'Renovated'
            WHEN 'UREN' THEN 'Renovated'
            WHEN 'CAN'  THEN 'Cancelled'
            WHEN 'NC'   THEN 'Never Completed'
            ELSE 'Other'
          END AS status_group,
          COUNT(*) AS count
        FROM   ctbuh_building b
        WHERE  b.deleted_at IS NULL
          AND  {filt_b_pie}
        GROUP  BY status_group
        ORDER  BY count DESC
    """)

    # ── 9. Height Distribution (bar) ──────────────────────────────────────────
    HEIGHT_ORDER = ['<100m', '100-149m', '150-199m', '200-299m',
                    '300-399m', '400-499m', '500m+']
    df_height = _bldg_query(f"""
        SELECT
          CASE
            WHEN b.height_architecture < 100 THEN '<100m'
            WHEN b.height_architecture < 150 THEN '100-149m'
            WHEN b.height_architecture < 200 THEN '150-199m'
            WHEN b.height_architecture < 300 THEN '200-299m'
            WHEN b.height_architecture < 400 THEN '300-399m'
            WHEN b.height_architecture < 500 THEN '400-499m'
            ELSE '500m+'
          END AS height_band,
          COUNT(*) AS count
        FROM   ctbuh_building b
        WHERE  b.deleted_at IS NULL
          AND  b.height_architecture IS NOT NULL
          AND  b.height_architecture > 0
          AND  {filt_b_pie}
        GROUP  BY height_band
        ORDER  BY MIN(b.height_architecture)
    """)

    STATUS_COLORS = {
        'Complete':           CVU_GREEN,
        'Under Construction': CVU_PALETTE[0],   # Indigo
        'Proposed':           CVU_PALETTE[1],   # Solar
        'Demolished':         CVU_PALETTE[5],   # Ember
        'Renovated':          CVU_PALETTE[2],   # Aqua
        'Cancelled':          CVU_PALETTE[4],   # Plum
        'Never Completed':    CVU_PALETTE[3],   # Teal
        'Other':              CVU_GRAY,
    }

    st_col, ht_col = st.columns(2)

    with st_col:
        if df_status.empty:
            st.info("No status data available.")
        else:
            df_st = df_status.sort_values('count', ascending=True)
            bar_colors = [STATUS_COLORS.get(s, CVU_GRAY) for s in df_st['status_group']]
            fig_st = go.Figure(go.Bar(
                x=df_st['count'],
                y=df_st['status_group'],
                orientation='h',
                marker_color=bar_colors,
                hovertemplate='<b>%{y}</b><br>%{x:,}<extra></extra>',
            ))
            fig_st.update_layout(**_chart_base(
                title=dict(text='Buildings by Status', font=dict(size=14, color=c['text'])),
                height=360,
                xaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                           tickfont=dict(color=c['subtext'], size=11), title=''),
                yaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                           tickfont=dict(color=c['subtext'], size=11),
                           title='', automargin=True),
                showlegend=False,
            ))
            st.plotly_chart(fig_st, use_container_width=True,
                            config={'displayModeBar': False})

    with ht_col:
        if df_height.empty:
            st.info("No height data available.")
        else:
            df_ht = df_height.copy()
            df_ht['height_band'] = pd.Categorical(
                df_ht['height_band'], categories=HEIGHT_ORDER, ordered=True)
            df_ht = df_ht.sort_values('height_band')
            fig_ht = go.Figure(go.Bar(
                x=df_ht['height_band'],
                y=df_ht['count'],
                marker_color=CVU_GREEN,
                hovertemplate='<b>%{x}</b><br>%{y:,} buildings<extra></extra>',
            ))
            fig_ht.update_layout(**_chart_base(
                title=dict(text='Height Distribution', font=dict(size=14, color=c['text'])),
                height=360,
                xaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                           tickfont=dict(color=c['subtext'], size=11), title=''),
                yaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                           tickfont=dict(color=c['subtext'], size=11),
                           title='', tickformat=',d'),
                showlegend=False,
            ))
            st.plotly_chart(fig_ht, use_container_width=True,
                            config={'displayModeBar': False})

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3 — DATA QUALITY
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### Data Quality")

    # 56 public-facing fields tracked for completeness
    _TOTAL_FIELDS = 56
    _SCORE_EXPR = """
        (b2.name_intl IS NOT NULL AND b2.name_intl != '') +
        (b2.name_local IS NOT NULL AND b2.name_local != '') +
        (b2.name_aka IS NOT NULL AND b2.name_aka != '') +
        (b2.name_fka IS NOT NULL AND b2.name_fka != '') +
        (b2.complex_id IS NOT NULL AND b2.complex_id != 0) +
        (b2.address IS NOT NULL AND b2.address != '') +
        (b2.zip IS NOT NULL AND b2.zip != '') +
        (b2.latitude IS NOT NULL AND b2.latitude != 0) +
        (b2.longitude IS NOT NULL AND b2.longitude != 0) +
        (b2.structural_material IS NOT NULL AND b2.structural_material != '') +
        (b2.material_displayed IS NOT NULL AND b2.material_displayed != '') +
        (b2.height_architecture IS NOT NULL AND b2.height_architecture != 0) +
        (b2.height_observatory IS NOT NULL AND b2.height_observatory != 0) +
        (b2.height_floor IS NOT NULL AND b2.height_floor != 0) +
        (b2.height_roof IS NOT NULL AND b2.height_roof != 0) +
        (b2.height_tip IS NOT NULL AND b2.height_tip != 0) +
        (b2.height_helipad IS NOT NULL AND b2.height_helipad != 0) +
        (b2.floors_above IS NOT NULL AND b2.floors_above != 0) +
        (b2.floors_below IS NOT NULL AND b2.floors_below != 0) +
        (b2.status IS NOT NULL AND b2.status != '') +
        (b2.proposed IS NOT NULL AND b2.proposed != 0) +
        (b2.start IS NOT NULL AND b2.start != 0) +
        (b2.completed IS NOT NULL AND b2.completed != 0) +
        (b2.demolished IS NOT NULL AND b2.demolished != 0) +
        (b2.main_use_01 IS NOT NULL AND b2.main_use_01 != '') +
        (b2.main_use_02 IS NOT NULL AND b2.main_use_02 != '') +
        (b2.main_use_03 IS NOT NULL AND b2.main_use_03 != '') +
        (b2.main_use_04 IS NOT NULL AND b2.main_use_04 != '') +
        (b2.main_use_05 IS NOT NULL AND b2.main_use_05 != '') +
        (b2.retrofit_use_01 IS NOT NULL AND b2.retrofit_use_01 != '') +
        (b2.retrofit_use_02 IS NOT NULL AND b2.retrofit_use_02 != '') +
        (b2.retrofit_use_03 IS NOT NULL AND b2.retrofit_use_03 != '') +
        (b2.retrofit_use_04 IS NOT NULL AND b2.retrofit_use_04 != '') +
        (b2.retrofit_use_05 IS NOT NULL AND b2.retrofit_use_05 != '') +
        (b2.elevators IS NOT NULL AND b2.elevators != 0) +
        (b2.elevator_speed IS NOT NULL AND b2.elevator_speed != 0) +
        (b2.gross_floor_area IS NOT NULL AND b2.gross_floor_area != 0) +
        (b2.usuable_floor_area IS NOT NULL AND b2.usuable_floor_area != 0) +
        (b2.parking IS NOT NULL AND b2.parking != 0) +
        (b2.apartments IS NOT NULL AND b2.apartments != 0) +
        (b2.office_space IS NOT NULL AND b2.office_space != 0) +
        (b2.hotel_rooms IS NOT NULL AND b2.hotel_rooms != 0) +
        (b2.commercial IS NOT NULL AND b2.commercial != 0) +
        (b2.observatory IS NOT NULL AND b2.observatory != 'no') +
        (b2.landmark_status IS NOT NULL AND b2.landmark_status != 'none') +
        (b2.energy_label IS NOT NULL AND b2.energy_label != '') +
        (b2.recladding_year IS NOT NULL AND b2.recladding_year != 0) +
        (b2.project_area IS NOT NULL AND b2.project_area != 0) +
        (b2.retrofit_start IS NOT NULL AND b2.retrofit_start != 0) +
        (b2.retrofit_end IS NOT NULL AND b2.retrofit_end != 0) +
        (b2.about IS NOT NULL AND b2.about != '') +
        (b2.trivia IS NOT NULL AND b2.trivia != '') +
        (b2.city_id IS NOT NULL AND b2.city_id != '') +
        (b2.country_id IS NOT NULL AND b2.country_id != '') +
        (b2.region_id IS NOT NULL AND b2.region_id != '') +
        (b2.timber IS NOT NULL AND b2.timber != 0)
    """

    # Completeness distribution (no date filter — always whole database)
    df_comp_dist = _bldg_query(f"""
        SELECT
          CASE
            WHEN score_pct < 20 THEN '0-19%'
            WHEN score_pct < 40 THEN '20-39%'
            WHEN score_pct < 60 THEN '40-59%'
            WHEN score_pct < 80 THEN '60-79%'
            ELSE '80-100%'
          END AS bucket,
          COUNT(*) AS building_count
        FROM (
          SELECT ROUND(( {_SCORE_EXPR} ) / {_TOTAL_FIELDS} * 100) AS score_pct
          FROM   ctbuh_building b2
          WHERE  b2.deleted_at IS NULL
        ) sub
        GROUP  BY bucket
        ORDER  BY MIN(score_pct)
    """)

    df_avg_comp = _bldg_query(f"""
        SELECT ROUND(AVG(( {_SCORE_EXPR} ) / {_TOTAL_FIELDS} * 100), 1) AS avg_pct
        FROM   ctbuh_building b2
        WHERE  b2.deleted_at IS NULL
    """)

    # Missing critical fields (no date filter)
    df_missing = _bldg_query("""
        SELECT
          SUM(CASE WHEN b.address IS NULL OR b.address = ''
              THEN 1 ELSE 0 END)                                     AS missing_address,
          SUM(CASE WHEN b.latitude IS NULL OR b.latitude = 0
              THEN 1 ELSE 0 END)                                     AS missing_latitude,
          SUM(CASE WHEN b.longitude IS NULL OR b.longitude = 0
              THEN 1 ELSE 0 END)                                     AS missing_longitude,
          SUM(CASE WHEN b.structural_material IS NULL OR b.structural_material = ''
              THEN 1 ELSE 0 END)                                     AS missing_material,
          COUNT(*)                                                    AS total
        FROM ctbuh_building b
        WHERE b.deleted_at IS NULL
    """)

    dq_col1, dq_col2 = st.columns(2)

    with dq_col1:
        if df_comp_dist.empty:
            st.info("No completeness data available.")
        else:
            avg_pct = (float(df_avg_comp['avg_pct'].iloc[0])
                       if not df_avg_comp.empty else 0.0)
            BUCKET_ORDER = ['0-19%', '20-39%', '40-59%', '60-79%', '80-100%']
            df_cd = df_comp_dist.copy()
            df_cd['bucket'] = pd.Categorical(
                df_cd['bucket'], categories=BUCKET_ORDER, ordered=True)
            df_cd = df_cd.sort_values('bucket')
            fig_comp = go.Figure(go.Bar(
                x=df_cd['bucket'],
                y=df_cd['building_count'],
                marker_color=CVU_GREEN,
                hovertemplate='<b>%{x}</b><br>%{y:,} buildings<extra></extra>',
            ))
            fig_comp.update_layout(**_chart_base(
                title=dict(
                    text=f'Record Completeness (avg: {avg_pct}%)',
                    font=dict(size=14, color=c['text']),
                ),
                height=340,
                xaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                           tickfont=dict(color=c['subtext'], size=11), title=''),
                yaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                           tickfont=dict(color=c['subtext'], size=11),
                           title='Buildings', tickformat=',d'),
                showlegend=False,
            ))
            st.plotly_chart(fig_comp, use_container_width=True,
                            config={'displayModeBar': False})

    with dq_col2:
        if df_missing.empty:
            st.info("No missing-data figures available.")
        else:
            total_bldgs = int(df_missing['total'].iloc[0])
            if total_bldgs > 0:
                field_labels = ['Address', 'Latitude', 'Longitude', 'Structural Material']
                field_cols   = ['missing_address', 'missing_latitude',
                                'missing_longitude', 'missing_material']
                miss_counts  = [int(df_missing[c].iloc[0]) for c in field_cols]
                miss_pcts    = [round(n / total_bldgs * 100, 1) for n in miss_counts]
                # Sort ascending so longest bar is at top
                pairs = sorted(zip(miss_pcts, miss_counts, field_labels))
                miss_pcts, miss_counts, field_labels = map(list, zip(*pairs))
                fig_miss = go.Figure(go.Bar(
                    x=miss_pcts,
                    y=field_labels,
                    orientation='h',
                    marker_color=CVU_PALETTE[5],
                    text=[f'{p}%  ({n:,})' for p, n in zip(miss_pcts, miss_counts)],
                    textposition='outside',
                    textfont=dict(color=c['text'], size=11),
                    hovertemplate=(
                        '<b>%{y}</b><br>%{x:.1f}% missing'
                        ' (%{customdata:,} buildings)<extra></extra>'
                    ),
                    customdata=miss_counts,
                ))
                fig_miss.update_layout(**_chart_base(
                    title=dict(text='Missing Critical Fields',
                               font=dict(size=14, color=c['text'])),
                    height=340,
                    xaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                               tickfont=dict(color=c['subtext'], size=11),
                               title='% of buildings',
                               range=[0, max(miss_pcts) * 1.35]),
                    yaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                               tickfont=dict(color=c['subtext'], size=11),
                               title='', automargin=True),
                    showlegend=False,
                ))
                st.plotly_chart(fig_miss, use_container_width=True,
                                config={'displayModeBar': False})

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 4 — GEOGRAPHY
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("### Geography")

    def _hbar_geo(df, label_col, value_col, title):
        """Horizontal bar chart sized to fit all rows."""
        if df.empty:
            st.info(f"No data for: {title}")
            return
        df = df.sort_values(value_col, ascending=True)  # highest at top
        height = max(300, len(df) * 24 + 80)
        fig = go.Figure(go.Bar(
            x=df[value_col],
            y=df[label_col],
            orientation='h',
            marker_color=CVU_GREEN,
            hovertemplate='<b>%{y}</b><br>%{x:,} buildings<extra></extra>',
        ))
        fig.update_layout(**_chart_base(
            title=dict(text=title, font=dict(size=14, color=c['text'])),
            height=height,
            xaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                       tickfont=dict(color=c['subtext'], size=11), title=''),
            yaxis=dict(gridcolor=c['grid'], zerolinecolor=c['grid'],
                       tickfont=dict(color=c['subtext'], size=11),
                       title='', automargin=True),
            showlegend=False,
            margin=dict(l=10, r=60, t=36, b=30),
        ))
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

    geo_view = st.radio(
        "Show:",
        ["New buildings (date added)", "Updated buildings (date modified)"],
        horizontal=True,
        key="bldg_geo_view",
    )
    geo_filt  = filt_b_pie if geo_view == "New buildings (date added)" else filt_b_update
    geo_label = "New" if geo_view == "New buildings (date added)" else "Updated"

    geo_tab_region, geo_tab_country, geo_tab_city, geo_tab_agg = st.tabs(
        ["Regions", "Countries", "Cities", "Agglomerations"]
    )

    with geo_tab_region:
        df_regions = _bldg_query(f"""
            SELECT r.name, COUNT(b.id) AS count
            FROM   ctbuh_building b
            JOIN   v2_regions r ON b.region_id = r.id
            WHERE  b.deleted_at IS NULL
              AND  {geo_filt}
            GROUP  BY r.name
            ORDER  BY count DESC
        """)
        _hbar_geo(df_regions, 'name', 'count', f'{geo_label} Buildings by Region')

    with geo_tab_country:
        df_countries = _bldg_query(f"""
            SELECT c.name, COUNT(b.id) AS count
            FROM   ctbuh_building b
            JOIN   v2_countries c ON b.country_id = c.id
            WHERE  b.deleted_at IS NULL
              AND  {geo_filt}
            GROUP  BY c.name
            ORDER  BY count DESC
            LIMIT  30
        """)
        _hbar_geo(df_countries, 'name', 'count', f'Top 30 Countries — {geo_label} Buildings')

    with geo_tab_city:
        df_cities = _bldg_query(f"""
            SELECT ci.name, COUNT(b.id) AS count
            FROM   ctbuh_building b
            JOIN   v2_cities ci ON b.city_id = ci.id
            WHERE  b.deleted_at IS NULL
              AND  {geo_filt}
            GROUP  BY ci.name
            ORDER  BY count DESC
            LIMIT  25
        """)
        _hbar_geo(df_cities, 'name', 'count', f'Top 25 Cities — {geo_label} Buildings')

    with geo_tab_agg:
        df_agg = _bldg_query(f"""
            SELECT a.name, COUNT(b.id) AS count
            FROM   ctbuh_building b
            JOIN   v2_cities ci  ON b.city_id = ci.id
            JOIN   agglomerations a ON ci.agglomeration_id = a.id
            WHERE  b.deleted_at IS NULL
              AND  {geo_filt}
            GROUP  BY a.name
            ORDER  BY count DESC
            LIMIT  25
        """)
        _hbar_geo(df_agg, 'name', 'count', f'Top 25 Agglomerations — {geo_label} Buildings')


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR + MAIN APP
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── Theme — must be first so CSS is injected before any content ───────────
    # ── Chart colour mode — dark by default, matches CVU brand ───────────────
    # Streamlit's own theme (from config.toml) handles page bg / widget colours.
    # This flag only controls Plotly chart colours (axis text, grid, plot_bg).
    if "dark_mode" not in st.session_state:
        st.session_state["dark_mode"] = True

    # ── Auth gate ─────────────────────────────────────────────────────────────
    if not _check_auth():
        return

    role = st.session_state.get("role", "staff")

    # ══════════════════════════════════════════════════════════════════════════
    # ADMIN (C-LEVEL) ROUTE — Time KPIs or Financial KPIs
    # ══════════════════════════════════════════════════════════════════════════
    if role == "admin":
        with st.sidebar:
            st.title("RTL Dashboard")
            st.divider()
            kpi_mode = st.radio(
                "KPI Module",
                options=["Time KPIs", "Financial KPIs", "Content KPIs", "Database KPIs"],
                key="admin_kpi_mode",
            )
            st.divider()
            # ── Chart colour toggle ───────────────────────────────────────────
            chart_mode = st.radio(
                "Chart Colors",
                options=["Dark", "Light"],
                index=0 if st.session_state["dark_mode"] else 1,
                key="admin_chart_mode",
                horizontal=True,
            )
            st.session_state["dark_mode"] = (chart_mode == "Dark")
            st.divider()
            if st.button("Refresh Data", use_container_width=True):
                load_all.clear()
                load_finances.clear()
                st.rerun()
            st.divider()
            if st.button("Sign Out", use_container_width=True):
                st.session_state.clear()
                st.rerun()

        if kpi_mode == "Time KPIs":
            st.title("Team Overview — Time KPIs")
            view_team(load_all())
        elif kpi_mode == "Financial KPIs":
            st.title("Financial KPIs")
            view_financial_kpis()
        elif kpi_mode == "Content KPIs":
            st.title("Content KPIs")
            view_content_kpis()
        else:
            st.title("Database KPIs")
            view_building_kpis()
        return

    # ══════════════════════════════════════════════════════════════════════════
    # STAFF ROUTE
    # ══════════════════════════════════════════════════════════════════════════

    # ── Landing page (no person chosen yet) ───────────────────────────────────
    if not st.session_state.get("person"):
        view_landing()
        return

    person = st.session_state["person"]

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.title("RTL KPI Tracker")
        st.divider()

        view = st.radio(
            "View",
            options=["My Time", "Team Overview"],
            label_visibility="collapsed",
        )

        st.divider()

        # ── Chart colour toggle ───────────────────────────────────────────────
        chart_mode = st.radio(
            "Chart Colors",
            options=["Dark", "Light"],
            index=0 if st.session_state["dark_mode"] else 1,
            key="staff_chart_mode",
            horizontal=True,
        )
        st.session_state["dark_mode"] = (chart_mode == "Dark")

        st.divider()

        st.caption(f"Signed in as **{person}**")
        if st.button("Change User", use_container_width=True):
            st.session_state["person"] = None
            st.rerun()

        st.divider()

        if st.button("Refresh Data", use_container_width=True):
            load_all.clear()
            st.rerun()

    # ── Load data (cached) ────────────────────────────────────────────────────
    df = load_all()

    # ── Route to view ─────────────────────────────────────────────────────────
    if view == "My Time":
        st.title(f"My Time — {person}")
        tab1, tab2, tab3 = st.tabs(["Daily Entry", "Bulk Edit", "History"])
        with tab1:
            view_daily_entry(person)
        with tab2:
            view_bulk_edit(person)
        with tab3:
            view_history(person, df)

    else:
        st.title("Team Overview")
        view_team(df)


if __name__ == "__main__":
    main()
